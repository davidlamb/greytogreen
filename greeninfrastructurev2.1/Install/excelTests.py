import openpyxl
from openpyxl import load_workbook
fileName = r"N:\Misc Projects\Grey_to_GreenInfrastructure\DEMOGIM\tampa_newdev.xlsx"
#fileName = r"N:\Misc Projects\Grey_to_GreenInfrastructure\Python Tool\greeninfrastructurev1\Install\projectWorkbook.xlsx" #
wb = load_workbook(fileName)
ws = wb['DataDictionary']
for row in ws.iter_rows(row_offset=1):
    for cell in row:
        cell.value = None
        pass

rowIndx = ws.max_row
print rowIndx

#ws.cell(row = rowIndx, column = 1).value = "something 2"
#ws.cell(row = rowIndx, column = 2).value = 3

wb.save(fileName)