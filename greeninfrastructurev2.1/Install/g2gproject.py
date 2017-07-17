import arcpy
from sharedtools import sharedTools
import os
import shutil
import csv
from g2glogging import Logging
import datetime
import time

import jdcal
import openpyxl
from openpyxl import load_workbook
import re
import sys

"""Settings class will be the controller for the arcpy toolbox view. Stores all the bmps that can be listed in a view. Communicates with the 'model' or database class"""
class G2GProjError(Exception):
    pass

class LicenseError(Exception):
    pass

class CriteriaProcessingError(Exception):
    pass


class Project(object):
    TABLES = ["settings","filetypes", "arearesults","fileorg"]
    #Settings Keys, See the geodatabase Table settings
    PROTECTIVE_BUFFER_KEY = "Protective Buffer (Feet)"
    DRAINAGE_AREA_KEY = "Drainage Area in Acres"
    STEEP_SLOPE_KEY = "Steep Slope (%)"
    ELEVATION_UNIT_KEY = "Elevation Units"
    PERCENT_TREE_KEY = "Percent Tree Canopy Cutoff (Percentage)"
    PERCENT_IMP_KEY = "Percent Impervious Canopy Cutoff (Percentage)"
    LANDCOVER_PREDEFINED = {"Trees":40,"Impervious":20,"Wood Wetlands":91,"Emergent Wetlands":92,"Water":10}
    SPLIT_CHARACTER = "|"
    TREES_NAME = "2. Trees/Forests"
    RIPARIAN_NAME = "1. Riparian Areas"
    RECHARGE_NAME = "3. Recharge Zones"
    DRAINAGE_NAME = "4. Natural Drainage Pathways"
    SLOPE_NAME = "5. Steep Slope"
    PERVIOUS_NAME = "6. Pervious Areas"
    IMPERVIOUS_RESULT_RASTER = "impervious_surface_raster"
    IMPERVIOUS_RESULT_VECOTR = "impervious_surface_vector"
    TREES_RESULT_RASTER = "trees_result_raster"
    TREES_RESULT_VECTOR = "trees_result_vector"

    def __init__(self):
        self.ProjectFolder = ""
        self.ProjectName = ""
        self.ProjectDatabase = ""
        self.ProjectDocument = ""
        self.ProjectReportDocument = ""
        self.ProjectBoundsOrigPath = ""
        self.ProjectBoundsDBPath = ""
        self.CellSize = 30

        self.ProjectPlanarUnitsCode = sharedTools.FOOT_US
        self.ElevationUnitString = sharedTools.ELEVATION_UNITS.keys()[0]
        self.ElevationUnitCode = sharedTools.FOOT_US
        #self._files = FileCollection()
        self._proj_extent = []#xmn,ymn,xmx,ymx
        self.ProjectSpatialRefCode = 9003

        self.currentAreasTable = "areas"
        self.ProjectWorkbook = ""
        self.ProjectRooftops = None
        self.ProjectManaged = None

        self.ProjectGreenInfrastructureMap=""
        self.ProjectCreditsAreas=""
        self.GIFResultsTable="resultsOrg"
        #Tool names in the geodatabase domain ToolNames
        self._riparianTool = "RiparianAreas"
        self._treesTool = "Trees"
        self._gwTool = "Groundwater"
        self._slopesTool = "Slopes"
        self._DrainagewaysTool = "DrainageWays"
        self._perviousTool = "PerviousAreas"
        #resultnames
        self._riparianLayer = None
        self._treesLayer = None
        self._rechargeLayer = None
        self._drainageLayer = None
        self._slopeLayer = None
        self._perviousLayer = None
        self._imperviousLayer = None

        self._deleteList = []
        self._logging = None

        self.Version = .1


    def checkOut(self):
        try:
            print "Check Out License"
            if arcpy.CheckExtension("Spatial") == "Available":
                arcpy.CheckOutExtension("Spatial")
                return True
            else:
                raise LicenseError
        except LicenseError:
            arcpy.AddError("Spatial Analyst Extension is not Available")
            arcpy.AddMessage("Spatial Analyst Extension is not Available")
            print "Spatial Analyst Extension is not Available"
        except:
            print arcpy.GetMessages(2)

    def updatePaths(self,updatePathFolder):
        oldDB = ""
        arcpy.AddMessage("Checking changes in folder path...")
        if self.ProjectFolder!="":
            arcpy.AddMessage("Current Project folder: %s"%self.ProjectFolder)
            arcpy.AddMessage("Using project folder %s"%updatePathFolder)
            if self.ProjectFolder != updatePathFolder:
                arcpy.AddMessage("Project Folder does not match....updating path information")
                self.ProjectFolder = updatePathFolder
                if self.ProjectDatabase !="":
                    oldDB = self.ProjectDatabase
                    pdbDir = os.path.dirname(self.ProjectDatabase)
                    basenm = os.path.basename(self.ProjectDatabase)
                    self.ProjectDatabase = os.path.join(updatePathFolder,basenm)
                    with arcpy.da.UpdateCursor(self.getFileOrgTablePath(),["InputFileLoc"]) as uc:
                        for row in uc:
                            row[0]=str(row[0]).replace(oldDB,self.ProjectDatabase)
                            uc.updateRow(row)
                if self.ProjectDocument!="":
                    pdbDir = os.path.dirname(self.ProjectDocument)
                    basenm = os.path.basename(self.ProjectDocument)
                    self.ProjectDocument= os.path.join(updatePathFolder,basenm)
                if self.ProjectBoundsDBPath !="":
                    pdbDir = os.path.dirname(self.ProjectBoundsDBPath)
                    basenm = os.path.basename(self.ProjectBoundsDBPath)
                    self.ProjectBoundsDBPath= os.path.join(self.ProjectDatabase,basenm)






    def createProject(self):
        if self.ProjectFolder !="":
            current = os.path.dirname(os.path.realpath(__file__))
            unique_name = arcpy.CreateUniqueName(self.ProjectName+".gdb", self.ProjectFolder)
            olddb = current + "\\basedb.xml"
            unique_name = os.path.basename(unique_name)
            #self.ProjectDatabase = arcpy.Copy_management(olddb,unique_name,"Workspace")[0]
            self.ProjectDatabase = arcpy.CreateFileGDB_management(self.ProjectFolder,unique_name)[0]
            arcpy.AddMessage(self.ProjectDatabase)
            arcpy.ImportXMLWorkspaceDocument_management(self.ProjectDatabase,olddb)

            newName = arcpy.CreateUniqueName(self.ProjectName + ".mxd",self.ProjectFolder)
            oldname = current + "\\template.mxd"
            self.ProjectDocument = arcpy.Copy_management(oldname,newName,"MapDocument")[0]
            newName = arcpy.CreateUniqueName(self.ProjectName + "_report.mxd",self.ProjectFolder)
            oldname = current + "\\template_pdf_output.mxd"
            self.ProjectReportDocument = arcpy.Copy_management(oldname,newName,"MapDocument")[0]

            oldwbName = current + "\\GIS_Outcomes_For_BMP_Scenario_Analysis_Tool.xlsx" #current + "\\projectWorkbook.xlsx" #current + "\\greytogreen.xlsm"
            newwbName = arcpy.CreateUniqueName(self.ProjectName + "_gisdata.xlsx",self.ProjectFolder) #arcpy.CreateUniqueName(self.ProjectName + ".xlsx",self.ProjectFolder) #arcpy.CreateUniqueName(self.ProjectName + ".xlsm",self.ProjectFolder)
            shutil.copyfile(oldwbName,newwbName)
            self.ProjectWorkbook = newwbName
            oldswbName = current + "\\BMP_Scenario_Analysis_Tool.xlsm"  # current + "\\projectWorkbook.xlsx" #current + "\\greytogreen.xlsm"
            newswbName = arcpy.CreateUniqueName(self.ProjectName + "_scenario.xlsm",
                                               self.ProjectFolder)  # arcpy.CreateUniqueName(self.ProjectName + ".xlsx",self.ProjectFolder) #arcpy.CreateUniqueName(self.ProjectName + ".xlsm",self.ProjectFolder)
            shutil.copyfile(oldswbName, newswbName)
            projBndName = arcpy.CreateUniqueName(self.ProjectName+"_bnd", self.ProjectDatabase)
            if arcpy.Exists(self.ProjectBoundsOrigPath):
                self.ProjectBoundsDBPath = arcpy.CopyFeatures_management(self.ProjectBoundsOrigPath,projBndName)[0]
            else:
                arcpy.AddError("Project Boundary Featurelayer does not exist.")

            #lyr = arcpy.MakeFeatureLayer_management(self.ProjectBoundsDBPath,"Project Boundaries")[0]
            mxd = arcpy.mapping.MapDocument(self.ProjectDocument)
            df = arcpy.mapping.ListDataFrames(mxd)[0]

            projectGroup = arcpy.mapping.Layer(current+"\\Project Area.lyr")

            if projectGroup.isGroupLayer:
                arcpy.AddMessage("Group Layer")

            arcpy.mapping.AddLayer(df,projectGroup,"TOP")
            projectGroup = arcpy.mapping.ListLayers(mxd,"Project Area")[0]
            lyr = arcpy.mapping.Layer(current+"\\project_boundaries.lyr")
            lyr.replaceDataSource(self.ProjectDatabase,"FILEGDB_WORKSPACE",os.path.basename(self.ProjectBoundsDBPath))
            lyr.visible = False
            arcpy.AddMessage("Replaced Datasource")

            arcpy.mapping.AddLayerToGroup(df,projectGroup,lyr)

            mxd.activeView = df.name

            resultsGroup = arcpy.mapping.Layer(current+"\\Results.lyr")
            arcpy.mapping.AddLayer(df,resultsGroup,"BOTTOM")

            #workingGroup = arcpy.mapping.Layer(current+"\\Working Layers.lyr")
            #arcpy.mapping.AddLayer(df,workingGroup,"BOTTOM")

            baseGroup = arcpy.mapping.Layer(current+"\\Base Layers.lyr")
            arcpy.mapping.AddLayer(df,baseGroup,"BOTTOM")



            mxd.save()
            for df in arcpy.mapping.ListDataFrames(mxd):
                df.spatialReference = self.getSpatialReference()

            mxd.save()
            self._logging = Logging(self.ProjectFolder)
            return True
        else:
            raise G2GProjError("Missing project folder!")
        return False


    def createRooftopFC(self):
        value = self.createFCWithAcres("Rooftops")
        if value:
            self.ProjectRooftops = value

    def createManagedAreasFC(self):
        value = self.createFCWithAcres("Managed")
        if value:
            self.ProjectManaged = value


    def createFCWithAcres(self,name):
        if not arcpy.Exists(self.ProjectDatabase + "\\"+name):
            roofFC = arcpy.CreateFeatureclass_management(self.ProjectDatabase,name,"POLYGON",spatial_reference=self.getSpatialReference())[0]
            arcpy.AddField_management(roofFC,"areaAcres","Double")
            fileOrgTable = self.getFileOrgTablePath()
            existingList = self.getFileOrgList()
            if name not in existingList:
                with arcpy.da.InsertCursor(fileOrgTable,self.getFileOrgTableFields()) as inc:
                    inc.insertRow([name,roofFC,roofFC])
            return roofFC
        else:
            return None

    def createFCByType(self, fcName, fcType):
        fcName = fcName.replace(" ","_")
        arcpy.env.addOutputsToMap = 0
        if not arcpy.Exists(self.ProjectDatabase + "\\"+fcName):
            newFC= arcpy.CreateFeatureclass_management(self.ProjectDatabase,fcName,"POLYGON",spatial_reference=self.getSpatialReference())[0]
            fileOrgTable = self.getFileOrgTablePath()
            existingList = self.getFileOrgList()
            if fcName not in existingList:
                with arcpy.da.InsertCursor(fileOrgTable,self.getFileOrgTableFields()) as inc:
                    inc.insertRow([fcType,newFC,newFC])
            return newFC

    def getSpatialReference(self):
        """Returns the spatial reference class from the code"""
        return arcpy.SpatialReference(self.ProjectSpatialRefCode)

    def getFileTypesTablePath(self):
        return self.ProjectDatabase+"\\FileTypes"

    def getFileTypesResultsPath(self):
        return self.ProjectDatabase+"\\FileTypesResults"

    def getFileOrgTablePath(self):
        return self.ProjectDatabase+"\\FileOrg"

    def getLandcovTablePath(self):
        return self.ProjectDatabase+"\\landcovermatch"

    def getSettingsTablePath(self):
        return self.ProjectDatabase+"\\settings"

    def getFileOrgTableFields(self):
        return ["FileType","InputFileLoc","OrigFileLoc"]

    def getSettingsTableFields(self):
        return ["setting","value"]

    def getToolSettingsTablePath(self):
        return self.ProjectDatabase+"\\toolsettings"

    def getNHDCodesTablePath(self):
        return self.ProjectDatabase+"\\nhdcodes"

    def getLandcoverMatchTablePath(self):
        return self.ProjectDatabase+"\\landcovermatch"

    def getAreaResultsTablePath(self):
        return self.ProjectDatabase+"\\arearesults"

    def getBaseFilesDataset(self):
        return self.ProjectDatabase

    def getFileTypesDictionary(self):
        ft = {}
        with arcpy.da.SearchCursor(self.getFileTypesTablePath(),["FileName","Geometry","FieldCheck"]) as sc:
            for row in sc:
                ft[row[0]] = {"geom":row[1],"field":row[2]}
        return ft


    def getFileOrgList(self):
        fo=[]
        with arcpy.da.SearchCursor(self.getFileOrgTablePath(),["FileType","OrigFileLoc"]) as sc:
            for row in sc:
                fo.append([row[1],row[0]])
        return fo


    def getSettingsDictionary(self):
        settings = {}
        with arcpy.da.SearchCursor(self.getSettingsTablePath(),self.getSettingsTableFields()) as sc:
            for row in sc:
                settings[row[0]]=row[1]
        return settings

    def getlandcoverValuesToMatch(self):
        fo=[]
        with arcpy.da.SearchCursor(self.getLandcovTablePath(),["landcovercode","shortdesc"]) as sc:
            for row in sc:
                merge = "%s%s%s"%(row[0],Project.SPLIT_CHARACTER,row[1])
                fo.append(merge)
        return fo

    def getNHDFeatureTypes(self):
        lst = {}
        with arcpy.da.SearchCursor(self.getNHDCodesTablePath(),["FeatureType","Layer","Active"])as sc:
            for row in sc:
                if row[1] in lst.keys():
                    lst[row[1]].append([row[0],row[2]])
                else:
                    lst[row[1]] = [[row[0],row[2]]]

        return lst

    def getActiveNHDFeatureTypesCodeList(self):
        lst = []
        with arcpy.da.SearchCursor(self.getNHDCodesTablePath(),["FeatureType","Active"],where_clause="active = 1")as sc:
            for row in sc:
                code = row[0].split(self.SPLIT_CHARACTER)[0]
                lst.append(code)

        return lst
    def updateActiveNHDFeatureTypes(self,values):
        with arcpy.da.UpdateCursor(self.getNHDCodesTablePath(),["FeatureType","Layer","Active"]) as uc:
            for row in uc:
                if row[0] in values:
                    row[2]= 1
                    uc.updateRow(row)
                else:
                    row[2] = 0
                    uc.updateRow(row)
        return True


    def getLandcoverMatchCode(self,inputCode):
        with arcpy.da.SearchCursor(self.getLandcoverMatchTablePath(),["landcovercode","matchlandcover"])as sc:
            for row in sc:
                if row[0]==str(inputCode):
                    return [int(x) for x in row[1].split(self.SPLIT_CHARACTER)]


    def updateLandcoverMatchCode(self,inputCodesD):
        """input dictionary key is landcover code, value is a list of match codes"""
        with arcpy.da.UpdateCursor(self.getLandcovTablePath(),["landcovercode","matchlandcover"]) as uc:
            arcpy.AddMessage("Landcovermatch")
            for row in uc:
                try:
                    row[1]=Project.SPLIT_CHARACTER.join(inputCodesD[row[0]])
                    arcpy.AddMessage(row)
                    uc.updateRow(row)
                except:
                    arcpy.AddMessage("Missing Key:%s"%(row[1]))
        return True


    def reviseFileOrg(self, inputFileList,copy=True):
        """Add Files to the project geodatabase. Expect it to be a list of lists with the path of the file first
        then the filetype"""
        fileOrgTable = self.getFileOrgTablePath()
        existingList = self.getFileOrgList()
        repeatedValues = []
        deletedValues = []
        current = os.path.dirname(os.path.realpath(__file__))
        arcpy.AddMessage("Input")
        arcpy.AddMessage(inputFileList)
        arcpy.AddMessage("Existing")
        arcpy.AddMessage(existingList)
        for i,j in enumerate(inputFileList):
            if j in existingList:
                repeatedValues.append(i)
        arcpy.AddMessage("Repeated")
        arcpy.AddMessage(repeatedValues)
        for o_i,o_j in enumerate(existingList):
            if o_j not in inputFileList:
                deletedValues.append(existingList[o_i])
        arcpy.AddMessage("Deleted")
        arcpy.AddMessage(deletedValues)



        with arcpy.da.UpdateCursor(fileOrgTable,self.getFileOrgTableFields()) as uc:
            for row in uc:
                temp = [row[2],row[0]]
                arcpy.AddMessage(temp)
                if temp in deletedValues:
                    arcpy.AddMessage("in Deleted values")
                    try:
                        desc = arcpy.Describe(row[1])
                        dt = desc.dataType
                        mxd = arcpy.mapping.MapDocument(self.ProjectDocument)
                        df = arcpy.mapping.ListDataFrames(mxd)[0]
                        lyr = arcpy.mapping.ListLayers(mxd,row[0])[0]
                        if lyr:
                            arcpy.mapping.RemoveLayer(df,lyr)
                            mxd.save()
                        arcpy.Delete_management(row[1],dt)
                    except:
                        arcpy.AddMessage("Error deleting the file and removing the layer...")
                    uc.deleteRow()
        #arcpy.DeleteRows_management(fileOrgTable)
        arcpy.env.workspace = self.ProjectFolder

        with arcpy.da.InsertCursor(fileOrgTable,self.getFileOrgTableFields()) as inc:
            for ind,val in enumerate( inputFileList):
                if ind not in repeatedValues:
                    desc = arcpy.Describe(val[0])
                    bn = desc.baseName
                    newDBPath = arcpy.CreateUniqueName(bn,self.ProjectDatabase)

                    if sharedTools.isRaster(val[0]):
                        mxd = arcpy.mapping.MapDocument(self.ProjectDocument)
                        df = arcpy.mapping.ListDataFrames(mxd)[0]
                        bndDesc = arcpy.Describe(self.ProjectBoundsDBPath)
                        expandFactor = float(.1) * bndDesc.extent.width
                        #newExtent = arcpy.Extent(bndDesc.extent.XMin-expandFactor,bndDesc.extent.YMin-expandFactor,
                                                 #bndDesc.extent.XMax+expandFactor,bndDesc.extent.YMax+expandFactor)
                        grpLayer = arcpy.mapping.ListLayers(mxd,"Base Layers")[0]
                        #newLoc = arcpy.Clip_management(val[0],str(newExtent),newDBPath)[0]
                        newLoc = arcpy.CopyRaster_management(val[0],newDBPath)[0]
                        #newLoc = arcpy.CopyRaster_management(val[0],newDBPath)[0]
                        lyrPath = os.path.join(current,"%s.lyr"%val[1])
                        if arcpy.Exists(lyrPath):
                            lyr = arcpy.mapping.Layer(lyrPath)
                            lyr.replaceDataSource(self.ProjectDatabase,"FILEGDB_WORKSPACE",os.path.basename(newLoc))
                        else:
                            lyr = arcpy.MakeRasterLayer_management(newLoc,val[1])[0]
                        lyr.visible=False
                        if grpLayer:
                            arcpy.mapping.AddLayerToGroup(df,grpLayer,lyr,"BOTTOM")
                        mxd.save()
                    else:
                        mxd = arcpy.mapping.MapDocument(self.ProjectDocument)
                        df = arcpy.mapping.ListDataFrames(mxd)[0]
                        grpLayer = arcpy.mapping.ListLayers(mxd,"Base Layers")[0]
                        newLoc = arcpy.Clip_analysis(val[0],self.ProjectBoundsDBPath,newDBPath)[0]
                        lyrPath = os.path.join(current,"%s.lyr"%val[1])
                        if arcpy.Exists(lyrPath):
                            lyr = arcpy.mapping.Layer(lyrPath)
                            lyr.replaceDataSource(self.ProjectDatabase,"FILEGDB_WORKSPACE",os.path.basename(newLoc))
                        else:
                            lyr = arcpy.MakeFeatureLayer_management(newLoc,val[1])[0]
                        lyr.visible=False
                        if grpLayer:
                            arcpy.mapping.AddLayerToGroup(df,grpLayer,lyr,"BOTTOM")
                        mxd.save()
                    inc.insertRow([val[1],newLoc,val[0]])

        mxd = arcpy.mapping.MapDocument(self.ProjectDocument)
        df = arcpy.mapping.ListDataFrames(mxd)[0]
        layerOrders =  self.getFileTypesLayerOrder()
        arcpy.AddMessage("Reordering Added Layers")
        grpLayer = arcpy.mapping.ListLayers(mxd,"Base Layers")[0]
        if grpLayer:
            for indx,val in enumerate(layerOrders):
                try:
                    currentLayer = arcpy.mapping.ListLayers(mxd,val[0])[0]
                    arcpy.mapping.RemoveLayer(df,currentLayer)
                    arcpy.mapping.AddLayerToGroup(df,grpLayer,currentLayer,"BOTTOM")#arcpy.mapping.MoveLayer(df, topLayer, currentLayer, "AFTER")
                except:
                    arcpy.AddMessage("error %s"%val[0])
            mxd.save()

    def resultsFileOrg(self, inputFileList,replace=True):
        """Add Files to the project geodatabase. Expect it to be a list of lists with the path of the file first
        then the filetype"""

        fileOrgTable = self.getFileOrgTablePath()
        existingList = self.getFileOrgList()
        repeatedValues = []
        deletedValues = []
        current = os.path.dirname(os.path.realpath(__file__))
        #arcpy.AddMessage("Input")
        #arcpy.AddMessage(inputFileList)
        #arcpy.AddMessage("Existing")
        #arcpy.AddMessage(existingList)
        for i,j in enumerate(inputFileList):
            for p,t in existingList:
                if j[1] == t:
                    repeatedValues.append(i)
        #arcpy.AddMessage("Repeated")
        #arcpy.AddMessage(repeatedValues)
        arcpy.env.workspace = self.ProjectFolder
        mxd = arcpy.mapping.MapDocument(self.ProjectDocument)
        df = arcpy.mapping.ListDataFrames(mxd)[0]
        resultsGrpLyr = arcpy.mapping.ListLayers(mxd,"Results")[0]
        arcpy.AddMessage("Add Results to FileOrg Table....")
        for ind,val in enumerate(inputFileList):
            #val = [path to file, filetype]
            if ind in repeatedValues:
                if replace:
                    #["FileType","InputFileLoc","OrigFileLoc"]
                    with arcpy.da.UpdateCursor(fileOrgTable,self.getFileOrgTableFields(),where_clause="FileType = '%s'"%val[1]) as uc:
                        for row in uc:
                            row[1] = val[0]
                            row[2] = val[0]
                            uc.updateRow(row)
                    try:
                        arcpy.AddMessage("Updating layer %s"%val[1])
                        lyr = arcpy.mapping.ListLayers(mxd,val[1])[0]
                        lyr.replaceDataSource(self.ProjectDatabase,"FILEGDB_WORKSPACE",os.path.basename(val[0]))

                    except Exception, e:
                        try:
                            lyrPath = os.path.join(current,"%s.lyr"%val[1])
                            lyr = arcpy.mapping.Layer(lyrPath)
                            lyr.replaceDataSource(self.ProjectDatabase,"FILEGDB_WORKSPACE",os.path.basename(val[0]))
                            if resultsGrpLyr:
                                arcpy.mapping.AddLayerToGroup(df,resultsGrpLyr,lyr,"TOP")
                            else:
                                arcpy.mapping.AddLayer(df,lyr,"TOP")
                        except:
                            try:
                                if sharedTools.isRaster(val[0]):
                                    lyr = arcpy.MakeRasterLayer_management(val[0],val[1])[0]
                                else:
                                    lyr = arcpy.MakeFeatureLayer_management(val[0],val[1])[0]
                                lyr.transparency = 30
                                if resultsGrpLyr:
                                    arcpy.mapping.AddLayerToGroup(df,resultsGrpLyr,lyr,"TOP")
                                else:
                                    arcpy.mapping.AddLayer(df,lyr,"TOP")
                            except:
                                pass

            else:
                with arcpy.da.InsertCursor(fileOrgTable,self.getFileOrgTableFields()) as inc:
                    inc.insertRow([val[1],val[0],val[0]])
                    lyrPath = os.path.join(current,"%s.lyr"%val[1])
                    arcpy.AddMessage("Add layer: %s"%lyrPath)
                    if arcpy.Exists(lyrPath):
                        try:
                            lyr = arcpy.mapping.Layer(lyrPath)
                            lyr.replaceDataSource(self.ProjectDatabase,"FILEGDB_WORKSPACE",os.path.basename(val[0]))
                            if resultsGrpLyr:
                                arcpy.mapping.AddLayerToGroup(df,resultsGrpLyr,lyr,"TOP")
                            else:
                                arcpy.mapping.AddLayer(df,lyr,"TOP")
                        except:
                            pass
                    else:
                        try:
                            if sharedTools.isRaster(val[0]):
                                lyr = arcpy.MakeRasterLayer_management(val[0],val[1])[0]
                            else:
                                lyr = arcpy.MakeFeatureLayer_management(val[0],val[1])[0]
                            lyr.transparency = 30
                            if resultsGrpLyr:
                                arcpy.mapping.AddLayerToGroup(df,resultsGrpLyr,lyr,"TOP")
                            else:
                                arcpy.mapping.AddLayer(df,lyr,"TOP")
                        except:
                            pass
        mxd.save()
        mxd = arcpy.mapping.MapDocument(self.ProjectDocument)
        df = arcpy.mapping.ListDataFrames(mxd)[0]
        layerOrders =  self.getFileResultsLayerOrder()
        arcpy.AddMessage("Reordering Added Layers")
        resultsGrpLyr = arcpy.mapping.ListLayers(mxd,"Results")[0]
        if resultsGrpLyr:
            for indx,val in enumerate(layerOrders):
                try:
                    currentLayer = arcpy.mapping.ListLayers(mxd,val[0])[0]
                    arcpy.mapping.RemoveLayer(df,currentLayer)
                    arcpy.mapping.AddLayerToGroup(df,resultsGrpLyr,currentLayer,"BOTTOM")#arcpy.mapping.MoveLayer(df, topLayer, currentLayer, "AFTER")
                except:
                    #arcpy.AddMessage("error %s"%val[0])
                    pass
        mxd.save()

    def outputPDFMap(self):

        mxd = arcpy.mapping.MapDocument(self.ProjectReportDocument)
        df = arcpy.mapping.ListDataFrames(mxd)[0]
        current = os.path.dirname(os.path.realpath(__file__))
        existingLayers = arcpy.mapping.ListLayers(mxd,data_frame=df)
        if len(existingLayers)>0:
            for lyr in existingLayers:
                arcpy.mapping.RemoveLayer(df,lyr)

        layerOrders =  self.getFileTypesLayerOrder()
        layerOrders += self.getFileResultsLayerOrder()

        layerOrders.sort(key=lambda x: x[1])
        arcpy.AddMessage(layerOrders)

        for indx,val in enumerate(layerOrders):
            try:
                fileType = val[0]
                filePath = self.getFiles(fileType)[0]
                lyrPath = os.path.join(current,"%s.lyr"%val[0])
                if sharedTools.isRaster(filePath) == False:
                    if arcpy.Exists(lyrPath):
                        lyr = arcpy.mapping.Layer(lyrPath)
                        arcpy.AddMessage(lyr.name)
                        lyr.replaceDataSource(self.ProjectDatabase,"FILEGDB_WORKSPACE",os.path.basename(filePath))
                        lyr.transparency = 0
                        arcpy.mapping.AddLayer(df,lyr,"BOTTOM")
                        arcpy.AddMessage("Add layer: %s"%lyrPath)
                    else:
                        if "Result" in fileType:
                            lyr = arcpy.MakeFeatureLayer_management(filePath,fileType)[0]
                            arcpy.mapping.AddLayer(df,lyr,"BOTTOM")
                            arcpy.AddMessage("Add layer: %s"%lyr.name)

            except:
                arcpy.AddMessage("error %s"%val[0])
        lyr = arcpy.mapping.Layer(current+"\\project_boundaries.lyr")
        lyr.replaceDataSource(self.ProjectDatabase,"FILEGDB_WORKSPACE",os.path.basename(self.ProjectBoundsDBPath))
        arcpy.mapping.AddLayer(df,lyr,"TOP")
        mxd.save()
        newName = arcpy.CreateUniqueName(self.ProjectName + "_report.pdf",self.ProjectFolder)
        arcpy.mapping.ExportToPDF(mxd, newName)


    def createGIMFC(self):
        arcpy.env.overwriteOutput=True
        self.ProjectGreenInfrastructureMap=arcpy.CreateFeatureclass_management(self.ProjectDatabase,"GreenInf","POLYGON",spatial_reference=self.getSpatialReference())[0]
        arcpy.AddField_management(self.ProjectGreenInfrastructureMap,"greeninfra","TEXT",field_length=100)
        arcpy.AddField_management(self.ProjectGreenInfrastructureMap,"areaAcres","DOUBLE")

    def createCreditsFeatClass(self):
        arcpy.env.overwriteOutput=True
        self.ProjectCreditsAreas=arcpy.CreateFeatureclass_management(self.ProjectDatabase,"GreenInfCredits","POLYGON",spatial_reference=self.getSpatialReference())[0]
        arcpy.AddField_management(self.ProjectCreditsAreas,"type","TEXT",field_length=100)
        arcpy.AddField_management(self.ProjectCreditsAreas,"areaAcres","DOUBLE")


    def clearValues(self):
        fileName = self.ProjectWorkbook #self.ProjectFolder + "\\" + self.ProjectName + ".xlsx"
        wb = load_workbook(fileName,read_only=False, keep_vba=True)
        ws = wb['DataDictionary']
        for row in ws.iter_rows(row_offset=1):
            for cell in row:
                cell.value=None
        wb.save(fileName)
        arcpy.CalculateField_management(in_table=self.getAreaResultsTablePath(), field="Value", expression="0", expression_type="VB", code_block="")

    def clearPreviousResults(self):
        fileResults = [x[0] for x in self.getFileResultsLayerOrder()]
        mxd = arcpy.mapping.MapDocument(self.ProjectDocument)
        df = arcpy.mapping.ListDataFrames(mxd)[0]
        #fileOrg = self.getFileOrgList()
        #["FileType","OrigFileLoc"]
        arcpy.AddMessage("Delete Previous Results")
        with arcpy.da.UpdateCursor(self.getFileOrgTablePath(),["FileType","OrigFileLoc"]) as uc:
            for row in uc:
                if row[0] in fileResults:

                    arcpy.AddMessage("Delete %s"%row[0])
                    arcpy.Delete_management(row[1])
                    for lyr in arcpy.mapping.ListLayers(mxd,row[0]):
                        arcpy.mapping.RemoveLayer(df,lyr)
                    mxd.save()
                    uc.deleteRow()





    def writeGIM(self):
        areas = {Project.RIPARIAN_NAME:{"mapped":"No","value":0.0,"units":"ACRES"},
         Project.TREES_NAME:{"mapped":"No","value":0.0,"units":"ACRES"},
         Project.RECHARGE_NAME:{"mapped":"No","value":0.0,"units":"ACRES"},
         Project.DRAINAGE_NAME:{"mapped":"No","value":0.0,"units":"ACRES"},
         Project.SLOPE_NAME:{"mapped":"No","value":0.0,"units":"ACRES"},
         Project.PERVIOUS_NAME:{"mapped":"No","value":0.0,"units":"ACRES"},
         "Project Boundary Area":{"mapped":"No","value":0.0,"units":"ACRES"}}

        #fileName = self.ProjectFolder + "\\GreenInfAreas.txt"
        fileName = self.ProjectWorkbook #self.ProjectFolder + "\\" + self.ProjectName + ".xlsm"
        arcpy.AddMessage(fileName)
        with arcpy.da.SearchCursor(self.ProjectGreenInfrastructureMap,["greeninfra","areaAcres"],sql_clause=(None,'ORDER BY greeninfra')) as sc:
            for row in sc:
                try:
                    areas[row[0]]["mapped"]="Yes"
                    if row[1]:
                        areas[row[0]]["value"] += row[1]
                        #areas[row[0]]["units"]="ACRES"
                except:
                    pass

        arcpy.env.workspace = self.ProjectDatabase
        arcpy.env.overwriteOutput=True

        cats = ["Protected Area", "Areas of Minimum Soil Compaction","Trees From Land Cover","Buildings"]
        buildingFile = ""
        protectedAreaFile = ""
        areasOfMinCompacFile = ""
        treesFile = None
        totalFound = 0
        for cat in cats:
            arcpy.AddMessage(cat)
            fileTypes = self.getFileTypesFromCat(cat)
            arcpy.AddMessage(fileTypes)
            for ft,geo,ub in fileTypes:
                eflst = self.getFiles(ft)
                if len(eflst)==1:
                    if cat == "Buildings":
                        buildingFile = eflst[0]
                        totalFound +=1
                        arcpy.AddMessage("Found Buildings")
                    if cat == "Protected Area":
                        protectedAreaFile = eflst[0]
                        #totalFound +=1
                        arcpy.AddMessage("Found Protected Area")
                    if cat == "Trees From Land Cover":
                        treesFile = eflst[0]
                        totalFound +=1
                        arcpy.AddMessage("Found Trees From Land Cover")
                    if cat == "Areas of Minimum Soil Compaction":
                        areasOfMinCompacFile = eflst[0]
                        totalFound +=1
                        arcpy.AddMessage("Areas of Minimum Soil Compaction")
        if arcpy.Exists(areasOfMinCompacFile):
            areas["Total Area of Minimum Soil Compaction"]={"mapped":"No","value":0.0,"units":""}
            areas["Total Area of Minimum Soil Compaction"]["value"] = self.calculateAcreAreaFromPolygons(areasOfMinCompacFile,"SQUAREFEET")
            areas["Total Area of Minimum Soil Compaction"]["units"]="SQUARE FEET"

        if arcpy.Exists(buildingFile):
            areas["Total Building Footprint Area"]={"mapped":"No","value":0.0,"units":""}
            areas["Total Building Footprint Area"]["value"] = self.calculateAcreAreaFromPolygons(buildingFile,"SQUAREFEET")
            areas["Total Building Footprint Area"]["units"]="SQUARE FEET"

        if arcpy.Exists(protectedAreaFile):
            areas["Total Protected Area"]={"mapped":"No","value":0.0,"units":""}
            areas["Total Protected Area"]["value"] = self.calculateAcreAreaFromPolygons(protectedAreaFile,"SQUAREFEET")
            areas["Total Protected Area"]["units"]="SQUARE FEET"

        if arcpy.Exists(treesFile) and arcpy.Exists(protectedAreaFile):
            protectedAreaGeom = arcpy.Dissolve_management(protectedAreaFile,arcpy.Geometry())[0]
            treesGeoms = arcpy.CopyFeatures_management(treesFile,arcpy.Geometry())
            if len(treesGeoms) >0:
                treesGeom=treesGeoms[0]
                treesGeom.difference(protectedAreaGeom.projectAs(treesGeom.spatialReference))
                areas["Total Tree Area Not in Protected Area"]={"mapped":"No","area":0.0,"units":""}
                areas["Total Tree Area Not in Protected Area"]["value"] = treesGeom.getArea("PLANAR","SQUAREFEET")
                areas["Total Tree Area Not in Protected Area"]["units"]="SQUARE FEET"

        areas["Project Boundary Area"]["value"]=self.calculateAcreAreaFromPolygons(self.ProjectBoundsDBPath)
        self.writeValuesToAreasTable(areas)

        #wb = load_workbook(fileName,read_only=False, keep_vba=True)
        #ws = wb['DataDictionary']
        #rowIndx = ws.max_row+1
        #for k in sorted(areas.keys()):
            #wrtr.writerow([k,areas[k]["mapped"],areas[k]["area"]])
            #ws.cell(row = rowIndx, column = 1).value = k
            #ws.cell(row=rowIndx, column=2).value = "{0:.2f}".format(areas[k]["area"])
            #ws.cell(row = rowIndx, column = 3).value = areas[k]["units"]
            #rowIndx +=1
        #wb.save(fileName)

    def writeValuesToAreasTable(self, valueDict):
        """Keys match values in the description column, values are areas in the units
        {Project.RIPARIAN_NAME:{"mapped":"No","value":0.0,"units":"ACRES"},
         Project.TREES_NAME:{"mapped":"No","value":0.0,"units":"ACRES"}}"""
        with arcpy.da.UpdateCursor(self.getAreaResultsTablePath(),["Description","Value"]) as uc:
            for row in uc:
                if row[0].strip() in valueDict.keys():
                    row[1]= valueDict[row[0].strip()]["value"]
                    uc.updateRow(row)
        return True

    def writeOutCurrentAreas(self):

        #if arcpy.Exists(self.ProjectWorkbook):
            #newWBName = arcpy.CreateUniqueName(self.ProjectName + "_gisdata.xlsx",self.ProjectFolder)
        #else:
            #newWBName = self.ProjectWorkbook
        arcpy.AddMessage("==============Exporting areas to Excel===============")
        #arcpy.AddMessage(newWBName)
        #try:
        #arcpy.TableToExcel_conversion(self.getAreaResultsTablePath(),newWBName)
        #self.ProjectWorkbook = newWBName
        #return True
        #except:
            #arcpy.AddMessage("ERROR SAVING TO NEW EXCEL FILE")
            #return False
        fileName = self.ProjectWorkbook#self.ProjectFolder + "\\" + self.ProjectName + ".xlsx"
        wb = load_workbook(fileName)
        ws = wb['DataDictionary']
        rowIndx = ws.max_row+1

        with arcpy.da.SearchCursor(self.getAreaResultsTablePath(),["Description","Value","Units"]) as sc:
            for row in sc:
                ws.cell(row = rowIndx, column = 1).value = row[0]
                ws.cell(row = rowIndx, column = 2).value = row[1]
                ws.cell(row = rowIndx, column = 3).value = row[2]
                rowIndx +=1
        wb.save(fileName)



    def createAreasTable(self):
        if not self._logging:
            self._logging = Logging(projectPath=self._ProjectFolder)
        self.clearPreviousResults()
        self.createGIMFC()
        arcpy.DeleteFeatures_management(self.ProjectGreenInfrastructureMap)
        if arcpy.Exists(self.ProjectDatabase+"\\"+self.GIFResultsTable):
            arcpy.DeleteRows_management(self.ProjectDatabase+"\\"+self.GIFResultsTable)
        self._deleteList = []
        self.createRiparianAreas()
        self.createTreeAreas()
        self.groundwaterRecharge()
        try:
            self.drainageWays()
            pass
        except:
            arcpy.AddMessage("Error calcualting Natural Drainage Pathways. You may not have the Spatial Analyst extension available...")
        try:
            self.slopes()
            pass
        except:
            arcpy.AddMessage("Error calcualting Slopes. You may not have the Spatial Analyst extension available...")
        self.perviousAreas()
        arcpy.AddMessage("================Calculate Areas=============")
        arcpy.CalculateField_management(self.ProjectGreenInfrastructureMap,"areaAcres","!shape.area@acres!","PYTHON_9.3","#")
        self.resultsFileOrg([[self.ProjectGreenInfrastructureMap,"Green Infrastructure Map"]])
        # mxd = arcpy.mapping.MapDocument(self.ProjectDocument)
        # df = arcpy.mapping.ListDataFrames(mxd)[0]
        # grpLayer = arcpy.mapping.ListLayers(mxd,"Results")[0]
        # #lyr = arcpy.MakeFeatureLayer_management(self.ProjectGreenInfrastructureMap,"Green Infrastructure Map")[0]
        # current = os.path.dirname(os.path.realpath(__file__))
        # lyr = arcpy.mapping.Layer(current+"\\Green Infrastructure Map Color.lyr")
        # lyr.replaceDataSource(self.ProjectDatabase,"FILEGDB_WORKSPACE",os.path.basename(self.ProjectGreenInfrastructureMap))
        # if grpLayer:
        #     arcpy.mapping.AddLayerToGroup(df,grpLayer,lyr,"TOP")
        # mxd.save()
        # tableName = arcpy.CreateUniqueName("areas",self.ProjectDatabase)
        # self.currentAreasTable = tableName
        # tableName = os.path.basename(tableName)
        # arcpy.CreateTable_management(self.ProjectDatabase,tableName)
        # arcpy.AddMessage(self.currentAreasTable)
        # arcpy.AddMessage("Adding Fields")
        # arcpy.AddField_management(self.currentAreasTable,"greeninfra","TEXT",field_length=100)
        # arcpy.AddField_management(self.currentAreasTable,"mapped","TEXT",field_length=4)
        # arcpy.AddField_management(self.currentAreasTable,"areaAcres","DOUBLE")


    def createCreditsTable(self):
        self.createCreditsFeatClass()
        arcpy.DeleteFeatures_management(self.ProjectCreditsAreas)
        self._deleteList = []
        self.treesWithinImpervious()
        arcpy.AddMessage("================Calculate Areas=============")
        arcpy.CalculateField_management(self.ProjectCreditsAreas,"areaAcres","!shape.area@acres!","PYTHON_9.3","#")

    def createResultsTable(self):
        pathToTable = self.ProjectDatabase+ "\\" + self.GIFResultsTable
        #if arcpy.Exists(pathToTable):
            #arcpy.DeleteRows_management(pathToTable)
        #else:
        arcpy.CreateTable_management(self.ProjectDatabase,self.GIFResultsTable)
        arcpy.AddField_management(self.GIFResultsTable,"type","TEXT",field_length=100)
        arcpy.AddField_management(self.GIFResultsTable,"path","TEXT",field_length=100)

    def addPathToResults(self,type,pathToFile):
        pathToTable = self.ProjectDatabase + "\\" + self.GIFResultsTable
        if not arcpy.Exists(pathToTable):
            self.createResultsTable()
        with arcpy.da.InsertCursor(pathToTable,["type","path"]) as ic:
            ic.insertRow([type,pathToFile])




    def cleanupDeleteList(self):
        for lyr in self._deleteList:
            try:
                arcpy.Delete_management(lyr)
            except:
                pass
        self._deleteList=[]

    def updateSettings(self, settingsDict):
        with arcpy.da.UpdateCursor(self.getSettingsTablePath(),self.getSettingsTableFields()) as uc:
            for row in uc:
                if row[0] in settingsDict.keys():
                    row[1]= settingsDict[row[0]]
                    uc.updateRow(row)
        return True

    def getRequiredFileCategories(self,toolName):
        lst=[]
        with arcpy.da.SearchCursor(self.getToolSettingsTablePath(),["tool","inputcategory"],where_clause="tool = '%s'"%(toolName)) as sc:
            for row in sc:
                lst.append(row[1])

        return lst

    def getFileTypesFromCat(self,category):
        lst=[]
        with arcpy.da.SearchCursor(self.getFileTypesTablePath(),["FileName","Geometry","User"],where_clause="Category = '%s'"%(category))as sc:
            for row in sc:
                lst.append([row[0],row[1],row[2]])
        with arcpy.da.SearchCursor(self.getFileTypesResultsPath(),["FileName","Geometry","User"],where_clause="Category = '%s'"%(category))as sc:
            for row in sc:
                lst.append([row[0],row[1],row[2]])
        return lst


    def getFileTypesLayerOrder(self):
        """Returns list of lists [FileName,LayerOrder]"""
        lst=[]
        with arcpy.da.SearchCursor(self.getFileTypesTablePath(),["FileName","LayerOrder"],sql_clause=(None,'ORDER BY LayerOrder'))as sc:
            for row in sc:
                lst.append([row[0],row[1]])
        return lst

    def getFileResultsLayerOrder(self):
        """Returns list of lists [FileName,LayerOrder]"""
        lst=[]
        with arcpy.da.SearchCursor(self.getFileTypesResultsPath(),["FileName","LayerOrder"],sql_clause=(None,'ORDER BY LayerOrder'))as sc:
            for row in sc:
                lst.append([row[0],row[1]])
        return lst

    def getFiles(self, fileType):
        lst = []
        with arcpy.da.SearchCursor(self.getFileOrgTablePath(),["InputFileLoc"],where_clause="FileType = '%s'"%(fileType))as sc:
            for row in sc:
                lst.append(row[0])

        return lst

    #def getImperviousAreas(self):


    def addToAreasFC(self,layerLocation,name):
        arcpy.AddMessage("Make sure geometry is correct...%s"%layerLocation)
        try:
            arcpy.RepairGeometry_management(layerLocation)
            GeomList = []
            with arcpy.da.SearchCursor(layerLocation, ["SHAPE@"]) as sc:
                for row in sc:
                    GeomList.append(row[0])

            if len(GeomList)>0:
                outputGeom=[]

                if arcpy.Exists(self.ProjectGreenInfrastructureMap):
                    with arcpy.da.SearchCursor(self.ProjectGreenInfrastructureMap,["SHAPE@"]) as sc:
                        for row in sc:
                            existingGeom = row[0]
                            for idx,Geom in enumerate(GeomList):
                                Geom = Geom.projectAs(self.getSpatialReference())
                                Geom = Geom.difference(existingGeom)
                                GeomList[idx] = Geom
                            arcpy.AddMessage("Subtracted out existing areas...")

                with arcpy.da.InsertCursor(self.ProjectGreenInfrastructureMap,["SHAPE@","greeninfra"]) as ic:
                    if len(GeomList) ==1:
                        ic.insertRow([GeomList[0],name])
                    else:
                        diss = arcpy.Dissolve_management(GeomList,arcpy.Geometry())
                        for Geom in diss:
                            ic.insertRow([Geom,name])

        except:
            arcpy.AddMessage("Failed to incorporate %s..."%name)
            arcpy.AddMessage("This could because %s has no geometry and is empty..."%name)


    def removeFromAreasFC(self,layerLocation):
        arcpy.RepairGeometry_management(layerLocation)
        Geom = None
        with arcpy.da.SearchCursor(layerLocation, ["SHAPE@"]) as sc:
            for row in sc:
                Geom = row[0]
                break
        if Geom:
            Geom = Geom.projectAs(self.getSpatialReference())

            if arcpy.Exists(self.ProjectGreenInfrastructureMap):
                with arcpy.da.UpdateCursor(self.ProjectGreenInfrastructureMap,["SHAPE@"]) as uc:
                    for row in uc:
                        try:
                            existingGeom = row[0]
                            existingGeom = existingGeom.difference(Geom)
                            row[0] = existingGeom
                            uc.updateRow(row)
                        except:
                            pass

#TODO clip the data to the project boundaries

    def createTreeAreas(self):
        arcpy.AddMessage("==============Trees==============")
        settings = self.getSettingsDictionary()
        protectiveBuffer = settings[self.PROTECTIVE_BUFFER_KEY] + " FEET"
        cats = self.getRequiredFileCategories(self._treesTool)
        arcpy.env.workspace = self.ProjectDatabase
        arcpy.env.overwriteOutput=True
        mergeLayers = []
        for cat in cats:
            arcpy.AddMessage(cat)
            arcpy.AddMessage(len(mergeLayers))
            if cat == "Trees":
                fileTypes = self.getFileTypesFromCat(cat)
                for ft,geo,ub in fileTypes:
                    vectorResult = None
                    if ub:
                        if geo == 'Polygon':
                            eflst = self.getFiles(ft)
                            if len(eflst) ==1:
                                arcpy.AddMessage("Tree cover is polygons defined by user...")
                                ef = eflst[0]
                                vectorResult = arcpy.MakeFeatureLayer_management(ef,"temp_trees_layer")[0]
                                if vectorResult:
                                    res = arcpy.MakeFeatureLayer_management(vectorResult,"trees_layer")[0]
                                    #res = arcpy.Buffer_analysis(res,"trees_buffer",protectiveBuffer,dissolve_option="ALL")[0]
                                    #self._deleteList.append(res)
                                    res = arcpy.Clip_analysis(res,self.ProjectBoundsDBPath,"trees_clip")[0]
                                    self._deleteList.append(res)
                                    lyr = arcpy.MakeFeatureLayer_management(res,"trees_layer")[0]
                                    mergeLayers.append(lyr)
                        if geo == "Raster":
                            eflst = self.getFiles(ft)
                            if len(eflst) ==1:
                                arcpy.AddMessage("Tree cover is raster, value of 1, defined by user....")
                                ef = eflst[0]
                                rast = arcpy.sa.Raster(ef)
                                trees = arcpy.sa.ExtractByAttributes(rast,"VALUE = 1")
                                if trees:
                                    res = arcpy.RasterToPolygon_conversion(trees,"treesuser_frmrast","NO_SIMPLIFY","VALUE")[0]
                                    self._deleteList.append(res)
                                    res = arcpy.Clip_analysis(res,self.ProjectBoundsDBPath,"treesuser_clip")[0]
                                    self._deleteList.append(res)
                                    lyr = arcpy.MakeFeatureLayer_management(res,"treesuser_layer")[0]
                                    mergeLayers.append(lyr)
                    else:
                        if geo == 'Raster':

                            #if it is raster, and not a user defined, it should be percentage.
                            percentTreeCanopy = float(settings[self.PERCENT_TREE_KEY])
                            eflst = self.getFiles(ft)
                            if len(eflst) ==1:
                                arcpy.AddMessage("Percent Tree Canopy...")
                                ef = eflst[0]
                                rast = arcpy.sa.Raster(ef)
                                res = arcpy.sa.Int(arcpy.sa.Con(rast,1,where_clause="VALUE>=%s"%percentTreeCanopy))
                                trees = arcpy.sa.ExtractByAttributes(res,"VALUE = 1")
                                trees.save("treesper_rast")
                                arcpy.Delete_management(res)
                                res = arcpy.RasterToPolygon_conversion(trees,"treesper_frmrast","NO_SIMPLIFY","VALUE")[0]
                                self._deleteList.append(res)
                                res = arcpy.Clip_analysis(res,self.ProjectBoundsDBPath,"treesper_clip")[0]
                                self._deleteList.append(res)
                                lyr = arcpy.MakeFeatureLayer_management(res,"treesper_layer")[0]
                                mergeLayers.append(lyr)
            if cat == "LandCover":
                fileTypes = self.getFileTypesFromCat(cat)
                vectorResult = None
                for ft,geo,ub in fileTypes:
                    if geo == "Raster":
                        if ub:

                            eflst = self.getFiles(ft)
                            if len(eflst) ==1:
                                arcpy.AddMessage("Using user-defined landcover raster")
                                arcpy.AddMessage("Matching landcover codes to predefined codes")
                                matchedCodes = self.getLandcoverMatchCode(self.LANDCOVER_PREDEFINED["Trees"]) + self.getLandcoverMatchCode(self.LANDCOVER_PREDEFINED["Wood Wetlands"])
                                arcpy.AddMessage(matchedCodes)
                                ef = eflst[0]
                                remapList = []
                                with arcpy.da.SearchCursor(ef, ["VALUE"]) as cursor:
                                    values = sorted({row[0] for row in cursor})
                                for row in values:
                                    if row in matchedCodes:
                                        remapList.append([row,1])
                                    else:
                                        remapList.append([row,0])
                                arcpy.AddMessage("Mapped Values")
                                arcpy.AddMessage(values)
                                arcpy.AddMessage(remapList)
                                if len(remapList)>0:
                                    remapValues = arcpy.sa.RemapValue(remapList)
                                    rast = arcpy.sa.Raster(ef)
                                    try:
                                        outReclass = arcpy.sa.Reclassify(rast, "VALUE", remapValues)
                                        #outReclass.save("trees_reclass")
                                        #self._deleteList.append("trees_reclass")
                                    except:
                                        m = "Error reclassifying: %s" %(cat)
                                        raise CriteriaProcessingError(m)
                                    res = arcpy.sa.ExtractByAttributes(outReclass,"VALUE = 1")
                                    #ext.save("trees_ext")
                                    #buffer the raster
                                    #bufferDist = float(settings[self.PROTECTIVE_BUFFER_KEY]) * sharedTools.getFactorFromFeet(ext.spatialReference.linearUnitCode)

                                    #dist = arcpy.sa.EucDistance(ext,bufferDist)
                                    #dist.save("trees_dist")
                                    #res = arcpy.sa.Int(arcpy.sa.Con(dist,1,where_clause="VALUE>=0"))
                                    #res.save("trees_dist_rcls")
                                    #arcpy.AddMessage("Save Trees Reclassify")
                                    #outReclass.save("trees_reclass")
                                    #self._deleteList.append("trees_reclass")
                                    arcpy.AddMessage("Trees to polygon")
                                    #vectorResult = arcpy.RasterToPolygon_conversion("trees_reclass","trees_frmrast","NO_SIMPLIFY","VALUE")[0]
                                    self._logging.write_output("Trees to polygon")
                                    vectorResult = arcpy.RasterToPolygon_conversion(res,"trees_frmrast","NO_SIMPLIFY","VALUE")[0]
                                    arcpy.AddMessage("Converted")
                                    self._deleteList.append(vectorResult)
                            if vectorResult:
                                arcpy.AddMessage("Feature Layer")
                                self._logging.write_output("Feature Layer")
                                res = arcpy.Clip_analysis(vectorResult,self.ProjectBoundsDBPath,"trees_clip")[0]
                                self._deleteList.append(res)
                                lyr = arcpy.MakeFeatureLayer_management(res,"trees_layer")[0]
                                mergeLayers.append(lyr)


            if cat == "Enviroatlas":
                fileTypes = self.getFileTypesFromCat(cat)
                vectorResult = None
                for ft,geo,ub in fileTypes:
                    if geo == "Raster":
                        eflst = self.getFiles(ft)
                        if len(eflst) ==1:
                            arcpy.AddMessage("Using Enviroatlas Dataset, codes for Forest and Woody Wetlands")
                            matchedCodes = [str(self.LANDCOVER_PREDEFINED["Trees"]),str(self.LANDCOVER_PREDEFINED["Wood Wetlands"])]
                            ef = eflst[0]
                            remapList = []
                            duplicateCheck = []
                            with arcpy.da.SearchCursor(ef,["VALUE"]) as sc:
                                for row in sc:
                                    if not row[0] in duplicateCheck:
                                        if str(row[0]) in matchedCodes:
                                            remapList.append([row[0],1])
                                            duplicateCheck.append(row[0])
                                        else:
                                            remapList.append([row[0],0])
                                            duplicateCheck.append(row[0])
                            arcpy.AddMessage("Remaplist")
                            arcpy.AddMessage(remapList)
                            self._logging.write_output("Remaplist")
                            self._logging.write_output(str(remapList))
                            if len(remapList)>0:
                                remapValues = arcpy.sa.RemapValue(remapList)
                                rast = arcpy.sa.Raster(ef)
                                arcpy.AddMessage("Reclassify")
                                self._logging.write_output("Reclassify")
                                try:
                                    outReclass = arcpy.sa.Reclassify(rast, "VALUE", remapValues)

                                except:
                                    m = "Error reclassifying: %s" %(cat)
                                    raise CriteriaProcessingError(m)
                                res = arcpy.sa.ExtractByAttributes(outReclass,"VALUE = 1")
                                #ext.save("trees_ext")
                                #buffer the raster
                                #bufferDist = float(settings[self.PROTECTIVE_BUFFER_KEY]) * sharedTools.getFactorFromFeet(ext.spatialReference.linearUnitCode)
                                #dist = arcpy.sa.EucDistance(ext,bufferDist)
                                #dist.save("trees_dist")
                                #res = arcpy.sa.Int(arcpy.sa.Con(dist,1,where_clause="VALUE>=0"))
                                #res.save("trees_dist_rcls")
                                #arcpy.AddMessage("Save Trees Reclassify")
                                #outReclass.save("trees_reclass")
                                #self._deleteList.append("trees_reclass")
                                arcpy.AddMessage("Trees to polygon")
                                #vectorResult = arcpy.RasterToPolygon_conversion("trees_reclass","trees_frmrast","NO_SIMPLIFY","VALUE")[0]
                                self._logging.write_output("Trees to polygon")
                                vectorResult = arcpy.RasterToPolygon_conversion(res,"trees_frmrast","NO_SIMPLIFY","VALUE")[0]
                                arcpy.AddMessage("Converted")
                                self._deleteList.append(vectorResult)
                            if vectorResult:
                                arcpy.AddMessage("Feature Layer")
                                self._logging.write_output("Feature Layer")
                                res = arcpy.Clip_analysis(vectorResult,self.ProjectBoundsDBPath,"trees_clip")[0]
                                self._deleteList.append(res)
                                lyr = arcpy.MakeFeatureLayer_management(res,"trees_layer")[0]
                                # arcpy.AddMessage("Buffer")
                                # self._logging.write_output("Buffer")
                                # res = arcpy.Buffer_analysis(res,"trees_buffer",protectiveBuffer,dissolve_option="ALL")[0]
                                # self._deleteList.append(res)
                                # self._logging.write_output("Clip")
                                # arcpy.AddMessage("Clip")
                                # res = arcpy.Clip_analysis(res,self.ProjectBoundsDBPath,"trees_clip")[0]
                                # self._deleteList.append(res)
                                # self._logging.write_output("Feature Layer")
                                # arcpy.AddMessage("Feature Layer")
                                # lyr = arcpy.MakeFeatureLayer_management(res,"trees_layer")[0]
                                mergeLayers.append(lyr)
        if len(mergeLayers)>0:
            arcpy.AddMessage(mergeLayers)
            arcpy.AddMessage("Union Layers")
            res=arcpy.Union_analysis(mergeLayers,"trees_union","NO_FID")[0]
            self._deleteList.append(res)
            arcpy.AddMessage("Dissolve")
            self._treesLayer = arcpy.Dissolve_management(res,"all_trees")[0]
            arcpy.AddMessage("DissolveComplete")
            self.addToAreasFC(self._treesLayer,self.TREES_NAME)
            arcpy.AddMessage("Write to results table")
            self.resultsFileOrg([[self._treesLayer,"Result Tree Areas"]])





    def createRiparianAreas(self):
        arcpy.AddMessage("==============Riparian Areas==============")
        settings = self.getSettingsDictionary()
        protectiveBuffer = settings[self.PROTECTIVE_BUFFER_KEY] + " FEET"
        cats = self.getRequiredFileCategories(self._riparianTool)
        arcpy.env.workspace = self.ProjectDatabase
        arcpy.env.overwriteOutput=True
        mergeLayers = []

        for cat in cats:
            if cat =="Wetland":
                arcpy.AddMessage("Wetlands")
                fileTypes = self.getFileTypesFromCat(cat)
                for ft,geo,ub in fileTypes:
                    if ub:
                        eflst = self.getFiles(ft)
                        if len(eflst) ==1:
                            arcpy.AddMessage("Wetlands - User defined polygons.")
                            ef = eflst[0]
                            res = arcpy.Buffer_analysis(ef,"wetlands_buffer",protectiveBuffer,dissolve_option="ALL")[0]
                            arcpy.AddMessage("Clean / Repair Geometry.")
                            arcpy.RepairGeometry_management(res)
                            self._deleteList.append(res)
                            res = arcpy.Clip_analysis(res,self.ProjectBoundsDBPath,"wetlands_clip")[0]
                            self._deleteList.append(res)
                            lyr = arcpy.MakeFeatureLayer_management(res,"wetlands_layer")[0]
                            mergeLayers.append(lyr)
                        elif len(eflst) > 1:
                            pass
            if cat == "Water" or cat=="Flow Lines":
                arcpy.AddMessage(cat)
                fileTypes = self.getFileTypesFromCat(cat)
                for ft,geo,ub in fileTypes:
                    if ub:
                        eflst = self.getFiles(ft)
                        if len(eflst) ==1:
                            arcpy.AddMessage("User-defined Water layers...")
                            ef = eflst[0]
                            uniqueNm = arcpy.CreateUniqueName("water_buffer",self.ProjectDatabase)
                            res = arcpy.Buffer_analysis(ef,uniqueNm,protectiveBuffer,dissolve_option="ALL")[0]
                            arcpy.AddMessage("Clean / Repair Geometry.")
                            arcpy.RepairGeometry_management(res)
                            self._deleteList.append(res)
                            res = arcpy.Clip_analysis(res,self.ProjectBoundsDBPath,uniqueNm+"_clip")[0]
                            self._deleteList.append(res)
                            lyr = arcpy.MakeFeatureLayer_management(res,uniqueNm+"_layer")[0]
                            mergeLayers.append(lyr)
                        elif len(eflst) > 1:
                            pass
                    else:
                        eflst = self.getFiles(ft)
                        if len(eflst) ==1:
                            arcpy.AddMessage("NHD layer, match feature type...")
                            lst = self.getActiveNHDFeatureTypesCodeList()
                            exp = "(" + ",".join(lst) + ")"
                            wc = 'FTYPE in '+exp
                            arcpy.AddMessage(wc)
                            ef = eflst[0]
                            uniqueNm = arcpy.CreateUniqueName("water_buffer",self.ProjectDatabase)
                            res = arcpy.MakeFeatureLayer_management(ef,"temporary_layer",wc)[0]
                            res = arcpy.Buffer_analysis(res,uniqueNm,protectiveBuffer,dissolve_option="ALL")[0]
                            arcpy.AddMessage("Clean / Repair Geometry.")
                            arcpy.RepairGeometry_management(res)
                            self._deleteList.append(res)
                            res = arcpy.Clip_analysis(res,self.ProjectBoundsDBPath,uniqueNm+"_clip")[0]
                            self._deleteList.append(res)
                            lyr = arcpy.MakeFeatureLayer_management(res,uniqueNm+"_layer")[0]
                            mergeLayers.append(lyr)
                            arcpy.Delete_management("temporary_layer")


            # if cat == "Soil":
            #     arcpy.AddMessage("Soil")
            #     fileTypes = self.getFileTypesFromCat(cat)
            #     wc = "hydgrp like 'C%' OR hydgrp = 'D'"
            #     for ft,geo,ub in fileTypes:
            #         eflst = self.getFiles(ft)
            #         if len(eflst) ==1:
            #             ef = eflst[0]
            #             res = arcpy.MakeFeatureLayer_management(ef,"temporary_layer",wc)[0]
            #             res = arcpy.Clip_analysis(res,self.ProjectBoundsDBPath,"soils_clip")[0]
            #             self._deleteList.append(res)
            #             lyr = arcpy.MakeFeatureLayer_management(res,"soils_layer")[0]
            #             mergeLayers.append(lyr)
            #             arcpy.Delete_management("temporary_layer")
            #         elif len(eflst) > 1:
            #             pass
            if cat == "Flood":
                arcpy.AddMessage("Flood")
                fileTypes = self.getFileTypesFromCat(cat)
                for ft,geo,ub in fileTypes:
                    if ub:
                        eflst = self.getFiles(ft)
                        if len(eflst) ==1:
                            arcpy.AddMessage("User-defined Flood layers...")
                            ef = eflst[0]
                            res = arcpy.Buffer_analysis(ef,"flood_buffer",protectiveBuffer,dissolve_option="ALL")[0]
                            arcpy.AddMessage("Clean / Repair Geometry.")
                            arcpy.RepairGeometry_management(res)
                            self._deleteList.append(res)
                            res = arcpy.Clip_analysis(res,self.ProjectBoundsDBPath,"flood_clip")[0]
                            self._deleteList.append(res)
                            lyr = arcpy.MakeFeatureLayer_management(res,"flood_layer")[0]
                            mergeLayers.append(lyr)
                        elif len(eflst) > 1:
                            pass
                    else:
                        eflst = self.getFiles(ft)
                        if len(eflst) ==1:
                            arcpy.AddMessage("FEMA Flood layer...")
                            ef = eflst[0]
                            wc = "FLD_ZONE LIKE 'A%'"
                            res = arcpy.MakeFeatureLayer_management(ef,"temporary_layer",wc)[0]
                            res = arcpy.Buffer_analysis(res,"flood_buffer",protectiveBuffer,dissolve_option="ALL")[0]
                            arcpy.AddMessage("Clean / Repair Geometry.")
                            arcpy.RepairGeometry_management(res)
                            self._deleteList.append(res)
                            res = arcpy.Clip_analysis(res,self.ProjectBoundsDBPath,"flood_clip")[0]
                            self._deleteList.append(res)
                            lyr = arcpy.MakeFeatureLayer_management(res,"flood_layer")[0]
                            mergeLayers.append(lyr)
                            arcpy.Delete_management("temporary_layer")
                        elif len(eflst) > 1:
                            pass
            if cat == "LandCover" or cat == "Enviroatlas":
                fileTypes = self.getFileTypesFromCat(cat)
                vectorResult = None
                for ft,geo,ub in fileTypes:
                    if geo == "Raster":
                        eflst = self.getFiles(ft)
                        if len(eflst) ==1:
                            arcpy.AddMessage("Landcover: either enviroatlas or alternative. Finding Emergent and Woody wetlands.")
                            arcpy.AddMessage("Matching landcover codes to predefined codes")
                            matchedCodes = self.getLandcoverMatchCode(self.LANDCOVER_PREDEFINED["Emergent Wetlands"]) + self.getLandcoverMatchCode(self.LANDCOVER_PREDEFINED["Wood Wetlands"])
                            arcpy.AddMessage(matchedCodes)
                            ef = eflst[0]
                            remapList = []
                            with arcpy.da.SearchCursor(ef, ["VALUE"]) as cursor:
                                values = sorted({row[0] for row in cursor})
                            for row in values:
                                if row in matchedCodes:
                                    remapList.append([row,1])
                                else:
                                    remapList.append([row,0])
                            arcpy.AddMessage("Mapped Values")
                            arcpy.AddMessage(values)
                            arcpy.AddMessage(remapList)
                            if len(remapList)>0:
                                remapValues = arcpy.sa.RemapValue(remapList)
                                rast = arcpy.sa.Raster(ef)
                                try:
                                    outReclass = arcpy.sa.Reclassify(rast, "VALUE", remapValues)
                                    outReclass.save("wetlands_lc_reclass")
                                    self._deleteList.append("wetlands_lc_reclass")
                                except:
                                    m = "Error reclassifying: %s" %(cat)
                                    raise CriteriaProcessingError(m)
                                vectorResult = arcpy.RasterToPolygon_conversion("wetlands_lc_reclass","wetlands_lc_frmrast","NO_SIMPLIFY","VALUE")[0]
                                self._deleteList.append(vectorResult)
                                if vectorResult:
                                    res = arcpy.MakeFeatureLayer_management(vectorResult,"wetlands_lc_layer","gridcode=1")[0]
                                    res = arcpy.Buffer_analysis(res,"wetlands_lc_buffer",protectiveBuffer,dissolve_option="ALL")[0]
                                    arcpy.AddMessage("Clean / Repair Geometry.")
                                    arcpy.RepairGeometry_management(res)
                                    self._deleteList.append(res)
                                    res = arcpy.Clip_analysis(res,self.ProjectBoundsDBPath,"wetlands_lc_clip")[0]
                                    self._deleteList.append(res)
                                    lyr = arcpy.MakeFeatureLayer_management(res,"wetlands_lc_layer")[0]
                                    mergeLayers.append(lyr)
            if cat == "Riparian Area":
                fileTypes = self.getFileTypesFromCat(cat)
                vectorResult = None
                for ft,geo,ub in fileTypes:
                    if ub == True:
                        eflst = self.getFiles(ft)
                        if len(eflst) ==1:
                            arcpy.AddMessage("User-defined Riparian Area...")
                            arcpy.AddMessage(eflst)
                            ef = eflst[0]
                            res = arcpy.MakeFeatureLayer_management(ef,"temporary_rip_layer")[0]
                            res = arcpy.Clip_analysis(res,self.ProjectBoundsDBPath,"riparian_clip")[0]
                            self._deleteList.append(res)
                            lyr = arcpy.MakeFeatureLayer_management(res,"riparian_user_layer")[0]
                            mergeLayers.append(lyr)

        arcpy.AddMessage(mergeLayers)
        if len(mergeLayers)>0:
            arcpy.AddMessage("Union Layers")
            res = arcpy.Union_analysis(mergeLayers,"riparian_union","NO_FID")[0]
            arcpy.AddMessage("Clean / Repair Geometry.")
            arcpy.RepairGeometry_management(res)
            self._deleteList.append(res)
            arcpy.AddMessage("Dissolve")
            self._riparianLayer = arcpy.Dissolve_management(res,"All_Riparian")[0]
            arcpy.AddMessage(self._riparianLayer)
            arcpy.AddMessage("Clean / Repair Geometry.")
            arcpy.RepairGeometry_management(self._riparianLayer)
            arcpy.AddMessage(self._riparianLayer)
            self.addToAreasFC(self._riparianLayer,self.RIPARIAN_NAME)
            arcpy.AddMessage("Write to results table")
            self.resultsFileOrg([[self._riparianLayer,"Result Riparian Area"]])
            # arcpy.AddMessage("Calculate Areas")
            # arcpy.AddField_management("FINAL_Riparian","areaacre","DOUBLE")
            # arcpy.CalculateField_management("FINAL_Riparian","areaacre","!shape.area@acres!","PYTHON_9.3","#")
            # ra_area = 0
            # with arcpy.da.SearchCursor("FINAL_Riparian",["areaacre"]) as sc:
            #     for row in sc:
            #         ra_area = row[0]
            #         break
            # with arcpy.da.InsertCursor(self.currentAreasTable,["greeninfra","mapped","areaAcres"]) as ic:
            #     ic.insertRow(["1. Riparian Area","Yes",ra_area])


#Groundwater recharge zones:
#Definition: areas that fall within soil groups A and B
#Inputs: Soils A and B Hydrologic Soil Group:
#Input: SSURGO spatial and tabular data.
#Methods: preprocess tabular database (as required by the SSURGO). Join spatial and tabular data using MUKEY to find soil groups. Extract soils matching the soil groups. Do not Buffer these polygons.
#Outputs: Polygon layer of groundwater recharge zones.

#Natural Drainage Ways:
#Definition: flow lines and pathways that are derived from the natural topography of the study area. These may not be represented in the hydrography layer.
#Input: Digital Elevation Model
#Methods: User specifies drainage area (default set to 5 acres). Flow Direction, Flow Accumulation, Flow lines. Buffer flow lines by 100 feet.
#Output: Polygon layer of all natural drainage ways.

    def groundwaterRecharge(self):
        arcpy.AddMessage("==============Groundwater Recharge Zones==============")
        settings = self.getSettingsDictionary()
        protectiveBuffer = settings[self.PROTECTIVE_BUFFER_KEY]
        cats = self.getRequiredFileCategories(self._gwTool)
        arcpy.env.workspace = self.ProjectDatabase
        arcpy.env.overwriteOutput=True
        mergeLayers = []

        for cat in cats:
            if cat == "Soil":
                arcpy.AddMessage("Soil")
                fileTypes = self.getFileTypesFromCat(cat)
                for ft,geo,ub in fileTypes:
                    eflst = self.getFiles(ft)
                    if len(eflst) ==1:
                        arcpy.AddMessage("Soil layer with hydgrp field.")
                        wc = "hydgrp ='A' OR hydgrp = 'B'"
                        ef = eflst[0]
                        res = arcpy.MakeFeatureLayer_management(ef,"temporary_layer",wc)[0]
                        res = arcpy.Clip_analysis(res,self.ProjectBoundsDBPath,"recharge_clip")[0]
                        self._deleteList.append(res)
                        lyr = arcpy.MakeFeatureLayer_management(res,"recharge_layer")[0]
                        mergeLayers.append(lyr)
                        arcpy.Delete_management("temporary_layer")
                    elif len(eflst) > 1:
                        pass
        if len(mergeLayers)>0:
            arcpy.AddMessage("Union Layers")
            res = arcpy.Union_analysis(mergeLayers,"recharge_union","NO_FID")[0]
            self._deleteList.append(res)
            arcpy.AddMessage("Dissolve")
            self._rechargeLayer = arcpy.Dissolve_management(res,"All_Recharge")[0]
            self.addToAreasFC(self._rechargeLayer,self.RECHARGE_NAME)
            arcpy.AddMessage("Write to results table")
            self.resultsFileOrg([[self._rechargeLayer,"Result Groundwater Recharge"]])


    def drainageWays(self):

        self.checkOut()
        arcpy.AddMessage("==============Natural Drainage Ways==============")
        settings = self.getSettingsDictionary()
        protectiveBuffer = settings[self.PROTECTIVE_BUFFER_KEY]
        drainageArea = float(settings[self.DRAINAGE_AREA_KEY])
        cats = self.getRequiredFileCategories(self._DrainagewaysTool)
        arcpy.env.workspace = self.ProjectDatabase
        arcpy.env.overwriteOutput=True
        mergeLayers = []

        for cat in cats:
            if cat =="Elevation":
                fileTypes = self.getFileTypesFromCat(cat)
                for ft,geo,ub in fileTypes:
                    eflst = self.getFiles(ft)
                    if len(eflst) ==1:
                        arcpy.AddMessage("Elevation")
                        arcpy.env.workspace = self.ProjectDatabase
                        ef = eflst[0]
                        rast = arcpy.Raster(ef)
                        rast_fill = arcpy.sa.Fill(rast)
                        rast_direc = arcpy.sa.FlowDirection(rast_fill)
                        rast_accum = arcpy.sa.FlowAccumulation(rast_direc)
                        arcpy.AddMessage("Create drainage areas based on DEM...")
                        rast_basins = arcpy.sa.Basin(rast_direc)
                        arcpy.env.workspace = self.ProjectDatabase
                        basinPolygons = arcpy.RasterToPolygon_conversion(rast_basins,"basins_fromdem","NO_SIMPLIFY","VALUE")[0]
                        self.resultsFileOrg([[basinPolygons,"Result Basins Derived From DEM"]])

                        #rast_accum.save("accum_rast")
                        x_cell = rast_accum.meanCellHeight
                        y_cell = rast_accum.meanCellWidth
                        cell_sq = x_cell*y_cell
                        linunits = rast_accum.spatialReference.linearUnitName
                        cell_sq_acres = sharedTools.getAcres(cell_sq,linunits)
                        arcpy.AddMessage("Cell area in acres: "+str(cell_sq_acres))
                        if cell_sq_acres:
                            da_rast = rast_accum * cell_sq_acres
                            #da_rast.save("da_rast")

                            #minr = da_rast.minimum-1
                            #maxr = da_rast.maximum+1

                            #remapRangeValues = arcpy.sa.RemapRange([[minr,drainageArea-.00001,0],[drainageArea,maxr,1],["NODATA","NODATA",0]])

                            #outReclass = arcpy.sa.Reclassify(da_rast, "Value", remapRangeValues)
                            #ext = arcpy.sa.ExtractByAttributes(outReclass,"VALUE = 1")
                            arcpy.env.workspace = self.ProjectDatabase
                            ext = arcpy.sa.Con(da_rast,1,where_clause="VALUE>=%s"%drainageArea)
                            arcpy.AddMessage("Create flow lines based on DEM...")
                            arcpy.env.workspace = self.ProjectDatabase
                            flowLines = arcpy.sa.StreamToFeature(ext,rast_direc,"flowlines_frmdem")

                            self.resultsFileOrg([[self.ProjectDatabase + "\\" + "flowlines_frmdem","Result Flowlines From DEM"]])
                            bufferDist = float(settings[self.PROTECTIVE_BUFFER_KEY]) * sharedTools.getFactorFromFeet(ext.spatialReference.linearUnitCode)
                            dist = arcpy.sa.EucDistance(ext,bufferDist)
                            res = arcpy.sa.Int(arcpy.sa.Con(dist,1,where_clause="VALUE>=0"))
                            vectorResult = arcpy.RasterToPolygon_conversion(res,"natdrainage_frmrast","NO_SIMPLIFY","VALUE")[0]
                            arcpy.AddMessage("Converted")
                            self._deleteList.append(vectorResult)
                            if vectorResult:
                                arcpy.AddMessage("Feature Layer")
                                self._logging.write_output("Feature Layer")
                                res = arcpy.Clip_analysis(vectorResult,self.ProjectBoundsDBPath,"natdrain_clip")[0]
                                self._deleteList.append(res)
                                lyr = arcpy.MakeFeatureLayer_management(res,"natdrainage_layer")[0]
                                mergeLayers.append(lyr)
                                arcpy.Delete_management(rast_fill)
                                arcpy.Delete_management(rast_direc)
                                arcpy.Delete_management(rast_accum)
                    elif len(eflst) > 1:
                        pass
        if len(mergeLayers)>0:
            arcpy.env.workspace = self.ProjectDatabase
            arcpy.AddMessage("Union Layers")
            res=arcpy.Union_analysis(mergeLayers,"natdrainage_union","NO_FID")[0]
            self._deleteList.append(res)
            arcpy.AddMessage("Dissolve")
            self._drainageLayer = arcpy.Dissolve_management(res,"All_Drainage")[0]
            self.addToAreasFC(self._drainageLayer,self.DRAINAGE_NAME)
            arcpy.AddMessage("Write to results table")
            self.resultsFileOrg([[self._drainageLayer,"Result Natural Drainage Pathways"]])

#Steep Slopes:
#Definition: Slope percentage greater than 25.
#Input: Digital Elevation Model
#Methods: Calculate percent slope. Reclassify slope by >25% areas. Buffer by 100 ft.
#Output: Polygon layer of all areas with >25% slope.

    def slopes(self):

        self.checkOut()
        arcpy.AddMessage("==============Steep Slopes==============")
        settings = self.getSettingsDictionary()
        protectiveBuffer = settings[self.PROTECTIVE_BUFFER_KEY]
        steepSlope = float(settings[self.STEEP_SLOPE_KEY])
        elevCode = sharedTools.ELEVATION_UNITS[settings[self.ELEVATION_UNIT_KEY]]
        cats = self.getRequiredFileCategories(self._slopesTool)
        arcpy.env.workspace = self.ProjectDatabase
        arcpy.env.overwriteOutput=True
        mergeLayers = []

        for cat in cats:
            vectorResult = None
            if cat =="Elevation":
                fileTypes = self.getFileTypesFromCat(cat)
                for ft,geo,ub in fileTypes:
                    eflst = self.getFiles(ft)
                    if len(eflst) ==1:
                        arcpy.AddMessage("Elevation - convert to % slope")
                        arcpy.env.workspace = self.ProjectDatabase
                        ef = eflst[0]
                        sr = arcpy.Describe(ef).spatialReference
                        zF = sharedTools.zfactorConvertZtoLinear(sr.linearUnitCode,elevCode)
                        rast = arcpy.Raster(ef)
                        try:
                            rast_slope = arcpy.sa.Slope(rast,"PERCENT_RISE",zF)
                            rast_slope.save("slope_rast")
                            arcpy.AddMessage("Created Slope")

                        except:
                            arcpy.AddError("ooopsss")

                        #minr = rast_slope.minimum-1
                        #maxr = rast_slope.maximum+1

                       #remapRangeValues = arcpy.sa.RemapRange([[minr,steepSlope-.00001,0],[steepSlope,maxr,1],["NODATA","NODATA",0]])

                        #outReclass = arcpy.sa.Reclassify(rast_slope, "Value", remapRangeValues)
                        outReclass = arcpy.sa.Con(rast_slope,1,where_clause="VALUE >= %s"%steepSlope)
                        #outReclass.save("reclass_rast")
                        vectorResult = arcpy.RasterToPolygon_conversion(outReclass,"steepslope","NO_SIMPLIFY","VALUE")[0]
                        if vectorResult:
                            res = arcpy.MakeFeatureLayer_management(vectorResult,"steepslope_layer","gridcode=1")[0]
                            #res = arcpy.Buffer_analysis(res,"natdrainage_buffer",protectiveBuffer,dissolve_option="ALL")[0]
                            #self._deleteList.append(res)
                            res = arcpy.Clip_analysis(res,self.ProjectBoundsDBPath,"steepslope_clip")[0]
                            self._deleteList.append(res)
                            lyr = arcpy.MakeFeatureLayer_management(res,"steepslope_layer2")[0]
                            mergeLayers.append(lyr)
                    elif len(eflst) > 1:
                        pass
        if len(mergeLayers)>0:
            arcpy.AddMessage("Union Layers")
            res=arcpy.Union_analysis(mergeLayers,"slope_union","NO_FID")[0]
            self._deleteList.append(res)
            arcpy.AddMessage("Dissolve")
            self._slopeLayer = arcpy.Dissolve_management(res,"all_slope")[0]
            self.addToAreasFC(self._slopeLayer,self.SLOPE_NAME)
            arcpy.AddMessage("Write to results table")
            try:
                self.resultsFileOrg([[self._slopeLayer,"Result Slope"]])
            except:
                pass


#Pervious Areas:
#Definition: Areas within project area that do not fall within the Green Infrastructure polygons created above and are not impervious.
#Input: EnviroAtlas Land Cover?.  User defined dataset.
#Methods: Reclassify to exclude developed land. Do not buffer. Remove all green infrastructure areas.
#Output: Polygon layer of open spaces.
    def perviousAreas(self):
        arcpy.AddMessage("==============Pervious Areas==============")
        settings = self.getSettingsDictionary()
        protectiveBuffer = settings[self.PROTECTIVE_BUFFER_KEY] + " FEET"
        cats = self.getRequiredFileCategories(self._perviousTool)
        arcpy.env.workspace = self.ProjectDatabase
        arcpy.env.overwriteOutput=True
        mergeLayers = []
        imperviousLayer = None
        for cat in cats:
            arcpy.AddMessage(cat)
            vectorResult = None
            if cat=="LandCover" or cat == "Enviroatlas":
                fileTypes = self.getFileTypesFromCat(cat)
                arcpy.AddMessage(fileTypes)
                vectorResult = None
                for ft,geo,ub in fileTypes:
                    arcpy.AddMessage(geo)
                    if geo == "Raster":
                        eflst = self.getFiles(ft)
                        arcpy.AddMessage(len(eflst))
                        if len(eflst) ==1:
                            arcpy.AddMessage("Enviroatlas")
                            arcpy.AddMessage("Matching landcover codes to predefined codes")
                            matchedCodes = self.getLandcoverMatchCode(self.LANDCOVER_PREDEFINED["Impervious"])
                            arcpy.AddMessage(matchedCodes)
                            ef = eflst[0]
                            remapList = []
                            with arcpy.da.SearchCursor(ef, ["VALUE"]) as cursor:
                                values = sorted({row[0] for row in cursor})
                            for row in values:
                                if row in matchedCodes:
                                    remapList.append([row,1])
                                else:
                                    remapList.append([row,0])
                            arcpy.AddMessage("Mapped Values")
                            arcpy.AddMessage(values)
                            arcpy.AddMessage(remapList)
                            if len(remapList)>0:
                                remapValues = arcpy.sa.RemapValue(remapList)
                                rast = arcpy.sa.Raster(ef)
                                try:
                                    outReclass = arcpy.sa.Reclassify(rast, "VALUE", remapValues)
                                    outReclass.save("pervious_reclass")
                                    imperv = arcpy.sa.ExtractByAttributes(outReclass,"VALUE = 1")
                                    imperv.save("impervious_bin")
                                except:
                                    m = "Error reclassifying: %s" %(cat)
                                    raise CriteriaProcessingError(m)
                                desc = arcpy.Describe("impervious_bin")
                                self.addPathToResults(self.IMPERVIOUS_RESULT_RASTER,desc.catalogPath)
                                self._deleteList.append("pervious_reclass")

                                vectorResult = arcpy.RasterToPolygon_conversion("pervious_reclass","pervious_frmrast","NO_SIMPLIFY","VALUE")[0]
                                self._deleteList.append(vectorResult)
                                if vectorResult:
                                    imp = arcpy.MakeFeatureLayer_management(vectorResult,"impervious_layer","gridcode=1")[0]
                                    res = arcpy.Clip_analysis(imp,self.ProjectBoundsDBPath,"impervious_clip")[0]
                                    self._deleteList.append(res)
                                    imperviousLayer = arcpy.CopyFeatures_management(res,"impervious_features")[0]
                                    arcpy.Delete_management(imp)
                                    res = arcpy.MakeFeatureLayer_management(vectorResult,"pervious_layer","gridcode=0")[0]
                                    res = arcpy.Clip_analysis(res,self.ProjectBoundsDBPath,"pervious_clip")[0]
                                    self._deleteList.append(res)
                                    lyr = arcpy.MakeFeatureLayer_management(res,"pervious_layer")[0]
                                    mergeLayers.append(lyr)
            if cat == "Impervious":
                fileTypes = self.getFileTypesFromCat(cat)
                vectorResult = None
                #todo add percentage cutoff
                for ft,geo,ub in fileTypes:
                    if geo == "Raster":
                        eflst = self.getFiles(ft)
                        if len(eflst) ==1:
                            arcpy.AddMessage("Percent impervious...")
                            settings = self.getSettingsDictionary()
                            protectiveBuffer = settings[self.PROTECTIVE_BUFFER_KEY]
                            percentImpervious = float(settings[self.PERCENT_IMP_KEY])

                            ef = eflst[0]
                            rast = arcpy.sa.Raster(ef)
                            res = arcpy.sa.Int(arcpy.sa.Con(rast,1,where_clause="VALUE>=%s"%percentImpervious))
                            res.save("imperviousStep")
                            imperv = arcpy.sa.ExtractByAttributes(res,"VALUE = 1")
                            imperv.save("impervious_rast")
                            desc = arcpy.Describe("impervious_rast")
                            self.addPathToResults(self.IMPERVIOUS_RESULT_RASTER,desc.catalogPath)
                            arcpy.Delete_management(res)
                            res = arcpy.sa.Int(arcpy.sa.Con(rast,1,where_clause="VALUE<%s"%percentImpervious))
                            res.save("perviousStep")
                            perv = arcpy.sa.ExtractByAttributes(res,"VALUE = 1")
                            perv.save("pervious_rast")
                            arcpy.Delete_management(res)
                            res = arcpy.RasterToPolygon_conversion(imperv,"impervious_frmrast","NO_SIMPLIFY","VALUE")[0]
                            self._deleteList.append(res)
                            res = arcpy.Clip_analysis(res,self.ProjectBoundsDBPath,"impervious_clip")[0]
                            self._deleteList.append(res)
                            imperviousLayer = arcpy.CopyFeatures_management(res,"impervious_features")[0]
                            res = arcpy.RasterToPolygon_conversion(perv,"pervious_frmrast","NO_SIMPLIFY","VALUE")[0]
                            self._deleteList.append(res)
                            res = arcpy.Clip_analysis(res,self.ProjectBoundsDBPath,"pervious_clip")[0]
                            self._deleteList.append(res)
                            lyr = arcpy.MakeFeatureLayer_management(res,"pervious_layer")[0]
                            mergeLayers.append(lyr)

        if len(mergeLayers)>0:
            arcpy.AddMessage("Union Layers")
            res=arcpy.Union_analysis(mergeLayers,"pervious_union","NO_FID")[0]
            self._deleteList.append(res)
            arcpy.AddMessage("Dissolve")
            self._perviousLayer = arcpy.Dissolve_management(res,"all_pervious")[0]
            self.addToAreasFC(self._perviousLayer,self.PERVIOUS_NAME)
            arcpy.AddMessage("Write to results table")
            self.resultsFileOrg([[self._perviousLayer,"Result Pervious from Land Cover"]])
        if imperviousLayer:
            arcpy.AddMessage("Prepare Impervious Layer")
            arcpy.AddMessage("Dissolve")
            arcpy.env.workspace = self.ProjectDatabase
            self._imperviousLayer = arcpy.Dissolve_management(imperviousLayer,"all_impervious")[0]
            self.removeFromAreasFC(self._imperviousLayer)
            arcpy.AddMessage("Write to results table")
            self.resultsFileOrg([[self._imperviousLayer,"Result Impervious from Land Cover"]])

    def treesWithinImpervious(self):
        arcpy.AddMessage("==============Identify Trees Near Impervious Areas==============")
        settings = self.getSettingsDictionary()
        distances = [20,100]

        arcpy.env.workspace = self.ProjectDatabase
        arcpy.env.overwriteOutput=True
        pathToTable = self.ProjectDatabase + "\\" + self.GIFResultsTable
        imperviousLayers = []
        imperviousBuffer = {}
        treeMergeLayers = []
        if not arcpy.Exists(pathToTable):
            arcpy.AddMessage("Results from green infrastructure map are not available....")
        try:
            with arcpy.da.SearchCursor(pathToTable,["type","path"]) as sc:
                for row in sc:
                    if row[0]==self.IMPERVIOUS_RESULT_RASTER or row[0]==self.IMPERVIOUS_RESULT_VECOTR:
                        imperviousLayers.append(row[1])
                    if row[0]==self.TREES_RESULT_VECTOR:
                        arcpy.AddMessage(self.TREES_RESULT_VECTOR)
                        treeMergeLayers.append(row[1])
        except:
            pass

        if len(imperviousLayers)==0:
            arcpy.AddMessage("Results from green infrastructure map are not available....")
            self.perviousAreas()
            try:
                with arcpy.da.SearchCursor(pathToTable,["type","path"]) as sc:
                    for row in sc:
                        if row[0]==self.IMPERVIOUS_RESULT_RASTER:
                            imperviousLayers.append(row[1])
            except:
                arcpy.AddError("Unable to find layers for impervious areas. Has landcover or percent imprevious been added?")
            if len(imperviousLayers)==0:
                arcpy.AddError("Unable to find layers for impervious areas. Has landcover or percent imprevious been added?")
        if len(treeMergeLayers)==0:
            arcpy.AddMessage("Results from green infrastructure map are not available....")
            self.createTreeAreas()
            try:
                with arcpy.da.SearchCursor(pathToTable,["type","path"]) as sc:
                    for row in sc:
                        if row[0]==self.TREES_RESULT_VECTOR:
                            arcpy.AddMessage(self.TREES_RESULT_VECTOR)
                            treeMergeLayers.append(row[1])
            except:
                arcpy.AddError("Unable to find layers for tree areas. Has landcover or percent tree canopy been added?")
            if len(treeMergeLayers)==0:
                arcpy.AddError("Unable to find layers for tree areas. Has landcover or percent tree canopy been added?")


        arcpy.AddMessage("Preprocessed layers available...")
        if len(imperviousLayers) > 0:
            first = None
            arcpy.AddMessage("Merging multiple impervious layers...")
            for lyr in imperviousLayers:
                rast = arcpy.Raster(lyr)
                if first != None:
                    resultRast = resultRast + arcpy.sa.Con(arcpy.sa.IsNull(rast),0,rast)
                else:
                    resultRast = arcpy.sa.Con(arcpy.sa.IsNull(rast),0,rast)
                    first = 1

            combined=arcpy.sa.Con(resultRast,1,where_clause="VALUE > 0")
            arcpy.AddMessage("Calculating buffered impervious areas...")
            for d in distances:
                arcpy.AddMessage("Current distance: %s"%d)
                nm = "imperv_r%s"%d
                ext = arcpy.sa.ExtractByAttributes(combined,"VALUE = 1")
                bufferDist = float(d) * sharedTools.getFactorFromFeet(ext.spatialReference.linearUnitCode)
                dist = arcpy.sa.EucDistance(ext,bufferDist)
                #dist.save("trees_dist")
                res = arcpy.sa.Int(arcpy.sa.Con(dist,1,where_clause="VALUE>=0"))
                vectorResult = arcpy.RasterToPolygon_conversion(res,nm,"NO_SIMPLIFY","VALUE")[0]
                arcpy.Delete_management(res)
                self._deleteList.append(vectorResult)
                res = arcpy.Clip_analysis(vectorResult,self.ProjectBoundsDBPath,nm+"_clip")[0]
                imp = arcpy.MakeFeatureLayer_management(res,nm+"_lyr")[0]
                self._deleteList.append(res)
                imperviousBuffer[d]=imp


        arcpy.AddMessage(len(treeMergeLayers))
        if len(treeMergeLayers)==1:
            arcpy.AddMessage("Load preprocessed trees layer...")
            treesLayer = arcpy.MakeFeatureLayer_management(treeMergeLayers[0],"trees_all_layer")[0]

        #if treesLayer:

        with arcpy.da.InsertCursor(self.ProjectCreditsAreas,["SHAPE@","TYPE","areaAcres"]) as ic:
            for d in distances:
                outputLabel = "Trees within %s feet of impervious area"%d
                arcpy.AddMessage(outputLabel)
                arcpy.AddMessage(imperviousBuffer[d])
                arcpy.AddMessage(treesLayer)
                resultGeom = arcpy.Intersect_analysis([treesLayer,imperviousBuffer[d]],arcpy.Geometry())
                #resultGeom = arcpy.Dissolve_management(resultGeom,arcpy.Geometry())[0]
                #arcpy.AddMessage(resultGeom.area)
                #arcpy.AddMessage(resultGeom)
                dissolveResult = arcpy.Dissolve_management(resultGeom,arcpy.Geometry())[0]
                areaValue = dissolveResult.getArea("PLANAR","ACRES")
                arcpy.AddMessage(dissolveResult)
                #for g in resultGeom:
                    #ic.insertRow([g,outputLabel])
                ic.insertRow([dissolveResult,outputLabel,areaValue])



    def calculateCriteriaAreas(self,bmpList):
        #domainList = arcpy.da.ListDomains(self.ProjectDatabase)
        #bmpList = []
        gtltp = re.compile(r"(?P<symbol><+=?|>+=?)(?P<number>\d*\.?\d*)")
        soilgrp = re.compile(r"([a-dA-D])?[-\s]?([a-dA-D])?")
        settings = self.getSettingsDictionary()
        elevCode = sharedTools.ELEVATION_UNITS[settings[self.ELEVATION_UNIT_KEY]]
        arcpy.env.workspace = self.ProjectDatabase
        arcpy.env.overwriteOutput = True
        uniqueCriteriaTable = os.path.basename(arcpy.CreateUniqueName("criteriaareas",self.ProjectDatabase))
        arcpy.AddMessage("Create Table: %s"%uniqueCriteriaTable)
        arcpy.CreateTable_management(out_path=self.ProjectDatabase,out_name=uniqueCriteriaTable)
        arcpy.AddField_management(in_table=uniqueCriteriaTable,field_name="bmptype",field_type="TEXT")
        arcpy.AddField_management(in_table=uniqueCriteriaTable,field_name="bmparea",field_type="DOUBLE")
        #for d in domainList:
            #if d.name.lower() == "bmpname":
                #for val, desc in d.codedValues.iteritems():
                    #bmpList.append(val)
        if len(bmpList)>0:
            for bmp in bmpList:
                lyrs = []
                q = "bmptype ='%s'"%bmp
                print q
                with arcpy.da.SearchCursor(self.ProjectDatabase+"//bmpcriteria",["bmptype","bmpcriteria","value"],where_clause=q) as sc:
                    for row in sc:
                        q2 = "criterianame ='%s'"%row[1]
                        print q2
                        with arcpy.da.SearchCursor(self.ProjectDatabase+"//criteriafilematch",["criterianame","filetype"],where_clause=q2) as ft_sc:
                            for cat in ft_sc:
                                print cat
                                fts = self.getFileTypesFromCat(cat[1])
                                print fts
                                for ft,type,ub in fts:
                                    eflst = self.getFiles(ft)
                                    print eflst
                                    if len(eflst) ==1:
                                        if row[1] == "Drainage Area":
                                            match = gtltp.search(row[2])
                                            symb = match.group('symbol')
                                            numVal = match.group('number')
                                            arcpy.AddMessage("Drainage Area Criteria")
                                            arcpy.AddMessage("%s%s"%(symb,numVal))
                                            ef = eflst[0]
                                            rast = arcpy.Raster(ef)
                                            rast_fill = arcpy.sa.Fill(rast)
                                            rast_direc = arcpy.sa.FlowDirection(rast_fill)
                                            rast_accum = arcpy.sa.FlowAccumulation(rast_direc)
                                            #rast_accum.save("accum_rast")
                                            x_cell = rast_accum.meanCellHeight
                                            y_cell = rast_accum.meanCellWidth
                                            cell_sq = x_cell*y_cell
                                            linunits = rast_accum.spatialReference.linearUnitName
                                            cell_sq_acres = sharedTools.getAcres(cell_sq,linunits)
                                            arcpy.AddMessage("Cell area in acres: "+str(cell_sq_acres))
                                            da_rast = rast_accum * cell_sq_acres
                                            da_rast.save("da_rast")
                                            wc = "VALUE%s%s"%(symb,numVal)
                                            try:
                                                res = arcpy.sa.Int(arcpy.sa.Con(da_rast,1,where_clause=wc))
                                                fin = arcpy.sa.ExtractByAttributes(res,"VALUE = 1")
                                                vectorResult = arcpy.RasterToPolygon_conversion(fin,"drainagearea%s"%numVal,"NO_SIMPLIFY","VALUE")[0]
                                                if vectorResult:
                                                    res = arcpy.Clip_analysis(vectorResult,self.ProjectBoundsDBPath,"drainageareacrit_clip")[0]
                                                    self._deleteList.append(res)
                                                    lyr = arcpy.MakeFeatureLayer_management(res,"drainageareacrit_layer2")[0]
                                                    lyrs.append(lyr)
                                            except:
                                                arcpy.AddMessage("Drainage Area Error")
                                        if row[1] == "Drainage Slope":
                                            match = gtltp.search(row[2])
                                            symb = match.group('symbol')
                                            numVal = match.group('number')
                                            arcpy.AddMessage("Elevation - convert to % slope")
                                            arcpy.AddMessage("%s%s"%(symb,numVal))
                                            ef = eflst[0]
                                            sr = arcpy.Describe(ef).spatialReference
                                            zF = sharedTools.zfactorConvertZtoLinear(sr.linearUnitCode,elevCode)
                                            rast = arcpy.Raster(ef)
                                            try:
                                                rast_slope = arcpy.sa.Slope(rast,"PERCENT_RISE",zF)
                                                rast_slope.save("slope_rast")
                                                arcpy.AddMessage("Created Slope")
                                            except:
                                                arcpy.AddError("ooopsss")
                                            wc = "VALUE%s%s"%(symb,numVal)
                                            try:
                                                res = arcpy.sa.Int(arcpy.sa.Con(rast_slope,1,where_clause=wc))
                                                fin = arcpy.sa.ExtractByAttributes(res,"VALUE = 1")
                                                vectorResult = arcpy.RasterToPolygon_conversion(fin,"slope%s"%numVal,"NO_SIMPLIFY","VALUE")[0]
                                                if vectorResult:
                                                    res = arcpy.Clip_analysis(vectorResult,self.ProjectBoundsDBPath,"bmpslope_clip")[0]
                                                    self._deleteList.append(res)
                                                    lyr = arcpy.MakeFeatureLayer_management(res,"bmpslope_layer2")[0]
                                                    lyrs.append(lyr)
                                            except:
                                                arcpy.AddMessage("Drainage Slope Error")
                                        if row[1] == "Hydrologic Soil Group":
                                            match = soilgrp.search(row[2])
                                            ef = eflst[0]
                                            arcpy.AddMessage("Soil Criteria")
                                            grpStr = ["'%s'"%(x) for x in match.groups()]
                                            soilquery = "hydgrp in (%s)"%(",".join(grpStr))
                                            arcpy.AddMessage(soilquery)
                                            try:
                                                lyr = arcpy.MakeFeatureLayer_management(ef,"soil_matchgroups",where_clause=soilquery)[0]
                                                lyrs.append(lyr)
                                            except:
                                                arcpy.AddMessage("Hydrologic Soil Group Criteria Error")
                                        if row[1] == "Road Buffer":
                                            match = gtltp.search(row[2])
                                            symb = match.group('symbol')
                                            numVal = match.group('number')
                                            arcpy.AddMessage("Road Criteria")
                                            arcpy.AddMessage("%s%s"%(symb,numVal))
                                            ef = eflst[0]
                                            try:
                                                if symb == "<" or symb == "<=":
                                                    res = arcpy.Buffer_analysis(ef,"roadcriteriaBuffer","%s FEET"%(numVal),dissolve_option="ALL")[0]
                                                    lyr = arcpy.MakeFeatureLayer_management(res,"roadcriteria_layer2")[0]
                                                    lyrs.append(lyr)
                                                if symb == ">" or symb == ">=":
                                                    res = self.inverseBuffer(ef, numVal, "roadcriteriaBuffer",self.ProjectBoundsDBPath)
                                                    lyr = arcpy.MakeFeatureLayer_management(res,"roadcriteria_layer2")[0]
                                                    lyrs.append(lyr)
                                            except:
                                                arcpy.AddMessage("Road Criteria Error")
                                        if row[1] == "Stream Buffer":
                                            match = gtltp.search(row[2])
                                            symb = match.group('symbol')
                                            numVal = match.group('number')
                                            arcpy.AddMessage("Stream Criteria")
                                            arcpy.AddMessage("%s%s"%(symb,numVal))
                                            ef = eflst[0]
                                            try:
                                                if symb == "<" or symb == "<=":
                                                    res = arcpy.Buffer_analysis(ef,"streamcriteriaBuffer","%s FEET"%(numVal),dissolve_option="ALL")[0]
                                                    lyr = arcpy.MakeFeatureLayer_management(res,"streamcriteria_layer2")[0]
                                                    lyrs.append(lyr)
                                                if symb == ">" or symb == ">=":
                                                    res = self.inverseBuffer(ef, numVal, "streamcriteriaBuffer",
                                                                             self.ProjectBoundsDBPath)
                                                    lyr = arcpy.MakeFeatureLayer_management(res,"streamcriteria_layer2")[0]
                                                    lyrs.append(lyr)
                                            except:
                                                arcpy.AddMessage("Stream Criteria Error")
                                        if row[1] == "Building Buffer":
                                            match = gtltp.search(row[2])
                                            symb = match.group('symbol')
                                            numVal = match.group('number')
                                            arcpy.AddMessage("Building Criteria")
                                            arcpy.AddMessage("%s%s"%(symb,numVal))
                                            ef = eflst[0]
                                            try:
                                                if symb == "<" or symb == "<=":
                                                    res = arcpy.Buffer_analysis(ef,"buildingcriteriaBuffer","%s FEET"%(numVal),dissolve_option="ALL")[0]
                                                    lyr = arcpy.MakeFeatureLayer_management(res,"buildingcriteria_layer2")[0]
                                                    lyrs.append(lyr)
                                                if symb == ">" or symb == ">=":
                                                    res = self.inverseBuffer(ef, numVal, "buildingcriteriaBuffer",
                                                                             self.ProjectBoundsDBPath)
                                                    lyr = arcpy.MakeFeatureLayer_management(res,"buildingcriteria_layer2")[0]
                                                    lyrs.append(lyr)
                                            except:
                                                arcpy.AddMessage("Building Criteria Error")

                uniqueBMPName = arcpy.CreateUniqueName(bmp.replace(" ","_").replace("/","_"),self.ProjectDatabase)
                print lyrs
                if len(lyrs)>1:
                    arcpy.Intersect_analysis(lyrs,uniqueBMPName)
                elif len(lyrs)==1:
                    arcpy.CopyFeatures_management(lyrs[0],uniqueBMPName)
                #Todo add a try and catch statement
                arcpy.AddMessage(uniqueBMPName)
                #calculate Area
                arcpy.AddMessage("Add Field")
                arcpy.AddField_management(uniqueBMPName,"areaAcres","Double")
                arcpy.AddMessage("Calculate Area")
                arcpy.CalculateField_management(uniqueBMPName,"areaAcres","!shape.area@acres!","PYTHON_9.3","#")
                totalArea = 0.0
                with arcpy.da.SearchCursor(uniqueBMPName,["areaAcres"]) as sc:
                    for row in sc:
                        totalArea+=row[0]
                with arcpy.da.InsertCursor(uniqueCriteriaTable,["bmptype","bmparea"]) as ic:
                    ic.insertRow([bmp,totalArea])
            #now write to table
            self.writeSiteCriteriaAreas(uniqueCriteriaTable)





    def buildPolygonFromFCExtent(self,inFC):
        extent = arcpy.Describe(inFC).extent
       # Array to hold points
        array = arcpy.Array()
        # Create the bounding box
        array.add(extent.lowerLeft)
        array.add(extent.lowerRight)
        array.add(extent.upperRight)
        array.add(extent.upperLeft)
        # ensure the polygon is closed
        array.add(extent.lowerLeft)
        # Create the polygon object
        polygon = arcpy.Polygon(array,arcpy.Describe(inFC).spatialReference)
        array.removeAll()
        return polygon

    def perviousWithSoil(self):
        arcpy.AddMessage("==============Pervious Areas and Soil Groups==============")
        arcpy.env.workspace = self.ProjectDatabase
        arcpy.env.overwriteOutput=True
        settings = self.getSettingsDictionary()
        soilCat = "Soil"
        soilFileTypes = self.getFileTypesFromCat(soilCat)
        soilFile = None
        soilLayer = None
        for ft,geo,ub in soilFileTypes:
            arcpy.AddMessage(ft)
            eflst = self.getFiles(ft)
            if len(eflst)==1:
                soilFile = eflst[0]
        if soilFile == None:
            arcpy.AddMessage("Cannot proceed. No dataset with soil hydrologic group found...")
        else:
            soilLayer = arcpy.MakeFeatureLayer_management(soilFile,"soilgrouplayer",where_clause="hydgrp IS NOT Null")[0]
            perviousNoTreesmergeLayers = []
            treesmergeLayers = []
            cats = ["Landcover","Enviroatlas","Impervious","Trees"]
            for cat in cats:
                arcpy.AddMessage(cat)
                vectorResult = None
                if cat=="LandCover" or cat == "Enviroatlas":
                    fileTypes = self.getFileTypesFromCat(cat)
                    arcpy.AddMessage(fileTypes)
                    vectorResult = None
                    for ft,geo,ub in fileTypes:
                        arcpy.AddMessage(geo)
                        if geo == "Raster":
                            eflst = self.getFiles(ft)
                            arcpy.AddMessage(len(eflst))
                            if len(eflst) ==1:
                                arcpy.AddMessage("Enviroatlas and Landcover")
                                arcpy.AddMessage("Matching landcover codes to predefined codes")
                                imperviousCode = self.getLandcoverMatchCode(self.LANDCOVER_PREDEFINED["Impervious"])
                                waterCode = self.getLandcoverMatchCode(self.LANDCOVER_PREDEFINED["Water"])
                                forestCodes = self.getLandcoverMatchCode(self.LANDCOVER_PREDEFINED["Wood Wetlands"])+self.getLandcoverMatchCode(self.LANDCOVER_PREDEFINED["Trees"])
                                ef = eflst[0]
                                with arcpy.da.SearchCursor(ef, ["VALUE"]) as cursor:
                                    values = sorted({row[0] for row in cursor})
                                remapListPerviousNonTrees = []
                                remapListPerviousTrees = []

                                for row in values:
                                    if row in imperviousCode:
                                        remapListPerviousNonTrees.append([row,0])
                                        remapListPerviousTrees.append([row,0])
                                    elif row in waterCode:
                                        remapListPerviousNonTrees.append([row,0])
                                        remapListPerviousTrees.append([row,0])
                                    elif row in forestCodes:
                                        remapListPerviousTrees.append([row,1])
                                        remapListPerviousNonTrees.append([row,0])
                                    else:
                                        remapListPerviousTrees.append([row,0])
                                        remapListPerviousNonTrees.append([row,1])

                                arcpy.AddMessage("Mapped Values")
                                arcpy.AddMessage(values)
                                arcpy.AddMessage(remapListPerviousTrees)
                                arcpy.AddMessage(remapListPerviousNonTrees)
                                if len(remapListPerviousNonTrees)>0:
                                    remapValues = arcpy.sa.RemapValue(remapListPerviousNonTrees)
                                    rast = arcpy.sa.Raster(ef)
                                    try:
                                        outReclass = arcpy.sa.Reclassify(rast, "VALUE", remapValues)
                                        outReclass.save("perviousnotrees_reclass")
                                        perv = arcpy.sa.ExtractByAttributes(outReclass,"VALUE = 1")
                                        perv.save("pervious_notrees")
                                        #desc = arcpy.Describe("impervious_bin")
                                        #self.addPathToResults(self.IMPERVIOUS_RESULT_RASTER,desc.catalogPath)
                                        self._deleteList.append("perviousnotrees_reclass")
                                    except:
                                        m = "Error reclassifying: %s" %(cat)
                                        raise CriteriaProcessingError(m)
                                    vectorResult = arcpy.RasterToPolygon_conversion("pervious_notrees","pervious_notrees_frmrast","NO_SIMPLIFY","VALUE")[0]
                                    self._deleteList.append(vectorResult)
                                    if vectorResult:
                                        res = arcpy.MakeFeatureLayer_management(vectorResult,"pervious_notrees_layer")[0]
                                        res = arcpy.Clip_analysis(res,self.ProjectBoundsDBPath,"pervious_notrees_clip")[0]
                                        self._deleteList.append(res)
                                        lyr = arcpy.MakeFeatureLayer_management(res,"pervious_notrees_layer")[0]
                                        perviousNoTreesmergeLayers.append(lyr)
                                if len(remapListPerviousTrees)>0:
                                    remapValues = arcpy.sa.RemapValue(remapListPerviousTrees)
                                    rast = arcpy.sa.Raster(ef)
                                    try:
                                        outReclass = arcpy.sa.Reclassify(rast, "VALUE", remapValues)
                                        outReclass.save("pervioustrees_reclass")
                                        perv = arcpy.sa.ExtractByAttributes(outReclass,"VALUE = 1")
                                        perv.save("pervious_trees")
                                        #desc = arcpy.Describe("impervious_bin")
                                        #self.addPathToResults(self.IMPERVIOUS_RESULT_RASTER,desc.catalogPath)
                                        self._deleteList.append("pervioustrees_reclass")
                                    except:
                                        m = "Error reclassifying: %s" %(cat)
                                        raise CriteriaProcessingError(m)
                                    vectorResult = arcpy.RasterToPolygon_conversion("pervious_trees","pervious_trees_frmrast","NO_SIMPLIFY","VALUE")[0]
                                    self._deleteList.append(vectorResult)
                                    if vectorResult:
                                        res = arcpy.MakeFeatureLayer_management(vectorResult,"pervious_trees_layer")[0]
                                        res = arcpy.Clip_analysis(res,self.ProjectBoundsDBPath,"pervious_trees_clip")[0]
                                        self._deleteList.append(res)
                                        lyr = arcpy.MakeFeatureLayer_management(res,"pervious_notrees_layer")[0]
                                        treesmergeLayers.append(lyr)
                if cat == "Impervious":
                    fileTypes = self.getFileTypesFromCat(cat)
                    vectorResult = None
                    #todo add percentage cutoff
                    for ft,geo,ub in fileTypes:
                        if geo == "Raster":
                            eflst = self.getFiles(ft)
                            if len(eflst) ==1:
                                arcpy.AddMessage("Percent impervious...")
                                settings = self.getSettingsDictionary()
                                protectiveBuffer = settings[self.PROTECTIVE_BUFFER_KEY]
                                percentImpervious = float(settings[self.PERCENT_IMP_KEY])

                                ef = eflst[0]
                                rast = arcpy.sa.Raster(ef)
                                res = arcpy.sa.Int(arcpy.sa.Con(rast,1,where_clause="VALUE<%s"%percentImpervious))
                                res.save("perperviousStep")
                                perv = arcpy.sa.ExtractByAttributes(res,"VALUE = 1")
                                perv.save("perpervious_rast")
                                res = arcpy.RasterToPolygon_conversion(perv,"perpervious_notreesfrmrast","NO_SIMPLIFY","VALUE")[0]
                                self._deleteList.append(res)
                                res = arcpy.Clip_analysis(res,self.ProjectBoundsDBPath,"perpervious_notrees_clip")[0]
                                self._deleteList.append(res)
                                lyr = arcpy.MakeFeatureLayer_management(res,"perpervious_notrees_layer")[0]
                                perviousNoTreesmergeLayers.append(lyr)
                if cat == "Trees":
                    fileTypes = self.getFileTypesFromCat(cat)
                    for ft,geo,ub in fileTypes:
                        vectorResult = None
                        if ub:
                            if geo == 'Polygon':
                                eflst = self.getFiles(ft)
                                if len(eflst) ==1:
                                    arcpy.AddMessage("Tree cover is polygons defined by user...")
                                    ef = eflst[0]
                                    vectorResult = arcpy.MakeFeatureLayer_management(ef,"temp_trees_layer")[0]
                                    if vectorResult:
                                        res = arcpy.MakeFeatureLayer_management(vectorResult,"trees_layer")[0]
                                        #res = arcpy.Buffer_analysis(res,"trees_buffer",protectiveBuffer,dissolve_option="ALL")[0]
                                        #self._deleteList.append(res)
                                        res = arcpy.Clip_analysis(res,self.ProjectBoundsDBPath,"trees_clip")[0]
                                        self._deleteList.append(res)
                                        lyr = arcpy.MakeFeatureLayer_management(res,"trees_layer")[0]
                                        treesmergeLayers.append(lyr)
                            if geo == "Raster":
                                eflst = self.getFiles(ft)
                                if len(eflst) ==1:
                                    arcpy.AddMessage("Tree cover is raster, value of 1, defined by user....")
                                    ef = eflst[0]
                                    rast = arcpy.sa.Raster(ef)
                                    trees = arcpy.sa.ExtractByAttributes(rast,"VALUE = 1")
                                    if trees:
                                        res = arcpy.RasterToPolygon_conversion(trees,"treesuser_frmrast","NO_SIMPLIFY","VALUE")[0]
                                        self._deleteList.append(res)
                                        res = arcpy.Clip_analysis(res,self.ProjectBoundsDBPath,"treesuser_clip")[0]
                                        self._deleteList.append(res)
                                        lyr = arcpy.MakeFeatureLayer_management(res,"treesuser_layer")[0]
                                        treesmergeLayers.append(lyr)
                        else:
                            if geo == 'Raster':

                                #if it is raster, and not a user defined, it should be percentage.
                                percentTreeCanopy = float(settings[self.PERCENT_TREE_KEY])
                                eflst = self.getFiles(ft)
                                if len(eflst) ==1:
                                    arcpy.AddMessage("Percent Tree Canopy...")
                                    ef = eflst[0]
                                    rast = arcpy.sa.Raster(ef)
                                    res = arcpy.sa.Int(arcpy.sa.Con(rast,1,where_clause="VALUE>=%s"%percentTreeCanopy))
                                    trees = arcpy.sa.ExtractByAttributes(res,"VALUE = 1")
                                    trees.save("treesper_rast")
                                    arcpy.Delete_management(res)
                                    res = arcpy.RasterToPolygon_conversion(trees,"treesper_frmrast","NO_SIMPLIFY","VALUE")[0]
                                    self._deleteList.append(res)
                                    res = arcpy.Clip_analysis(res,self.ProjectBoundsDBPath,"treesper_clip")[0]
                                    self._deleteList.append(res)
                                    lyr = arcpy.MakeFeatureLayer_management(res,"treesper_layer")[0]
                                    treesmergeLayers.append(lyr)

            if len(treesmergeLayers)>0 and len(perviousNoTreesmergeLayers)>0:
                arcpy.AddMessage("Union")
                arcpy.AddMessage(treesmergeLayers)
                arcpy.AddMessage(perviousNoTreesmergeLayers)
                res_pervTrees=arcpy.Union_analysis(treesmergeLayers,"pervtrees_union","NO_FID")[0]
                res_pervNoTrees=arcpy.Union_analysis(perviousNoTreesmergeLayers,"pervnotrees_union","NO_FID")[0]
                #self._deleteList.append(res_pervTrees)
                #self._deleteList.append(res_pervNoTrees)
                arcpy.AddMessage("Get polygon geometry to check for overlap...")
                perviousNoTreesGeom = arcpy.Dissolve_management(res_pervNoTrees,arcpy.Geometry())[0]
                perviousTreesGeom = arcpy.Dissolve_management(res_pervTrees,arcpy.Geometry())[0]
                arcpy.AddMessage("Test overlapping geometry...")
                doesitoverlap = perviousNoTreesGeom.overlaps(perviousTreesGeom)
                arcpy.AddMessage("Overlap Result: %s"%(doesitoverlap))

                if doesitoverlap == True:
                    #perviousNoTreesGeom = arcpy.Dissolve_management(res_pervNoTrees,"pervnotrees_union")[0]
                    arcpy.AddMessage("Remove pervious trees from pervious areas...")
                    diff = perviousNoTreesGeom.difference(perviousTreesGeom)
                    res_pervNoTrees = arcpy.CopyFeatures_management(diff,"perviousNoTrees_Subtract")[0]

                #perviousNoTreesSoil = arcpy.Intersect_analysis(perviousNoTreesMinusTrees,"pervious_soils_intersect","ALL")
                arcpy.AddMessage("Pervious No Trees Intersect")
                #perviousNoTreesSoil = arcpy.Intersect_analysis([res_pervNoTrees,soilLayer],"pervious_soils_intersect","ALL")[0]
                perviousNoTreesSoil = arcpy.Intersect_analysis([res_pervNoTrees,soilLayer],"pervious_soils_intersect","ALL")[0]
                #perviousNoTreesSoilDissolve = arcpy.Dissolve_management(perviousNoTreesSoil,"all_pervious_soils")[0]
                arcpy.AddMessage("Pervious Trees Intersect")
                perviousTreesSoil = arcpy.Intersect_analysis([res_pervTrees,soilLayer],"trees_soils_intersect","ALL")[0]
                arcpy.AddMessage("Pervious No Trees Dissolve")
                allPerviousSoil = arcpy.Dissolve_management(perviousNoTreesSoil,"all_pervious_soil","hydgrp")[0]
                arcpy.AddMessage("Add Field")
                arcpy.AddField_management(allPerviousSoil,"areaAcres","Double")
                arcpy.AddMessage("Calculate Area")
                arcpy.CalculateField_management(allPerviousSoil,"areaAcres","!shape.area@acres!","PYTHON_9.3","#")
                arcpy.AddMessage("Pervious Trees Dissolve")
                allTreesSoil = arcpy.Dissolve_management(perviousTreesSoil,"all_trees_soil","hydgrp")[0]
                arcpy.AddMessage("Add Field")
                arcpy.AddField_management(allTreesSoil,"areaAcres","Double")
                arcpy.AddMessage("Calculate Area")
                arcpy.CalculateField_management(allTreesSoil,"areaAcres","!shape.area@acres!","PYTHON_9.3","#")
                #perviousTreesSoilDissolve = arcpy.Dissolve_management(perviousTreesSoil,"all_trees_soils")[0]

    def calculateCriteriaAreasNew(self):
        arcpy.AddMessage("==============Siting Criteria==============")
        arcpy.env.workspace = self.ProjectDatabase
        arcpy.env.overwriteOutput=True
        settings = self.getSettingsDictionary()
        elevCode = sharedTools.ELEVATION_UNITS[settings[self.ELEVATION_UNIT_KEY]]
        arcpy.AddMessage("Checking for input files...")
        cats = ["Buildings","Elevation","Parking Impervious","Impervious From Land Cover","Riparian Area", "Protected Area", "Areas of Minimum Soil Compaction"]


        buildingFile = ""
        demFile = ""
        proposedImperviousFile = ""
        existingImperviousFile = ""
        existingRiparianFile = ""
        protectedAreaFile = ""
        areasOfMinCompacFile = ""
        totalFound = 0
        resultsList = {}
        for cat in cats:
            fileTypes = self.getFileTypesFromCat(cat)
            for ft,geo,ub in fileTypes:
                eflst = self.getFiles(ft)
                if len(eflst)==1:
                    if cat == "Buildings":
                        buildingFile = eflst[0]
                        totalFound +=1
                        arcpy.AddMessage("Found Buildings")
                    if cat == "Elevation":
                        demFile = eflst[0]
                        totalFound +=1
                        arcpy.AddMessage("Found Elevation")
                    if cat == "Parking Impervious":
                        proposedImperviousFile = eflst[0]
                        totalFound +=1
                        arcpy.AddMessage("Found Parking Impervious")
                    if cat == "Impervious From Land Cover":
                        existingImperviousFile = eflst[0]
                        totalFound +=1
                        arcpy.AddMessage("Found Impervious from Land Cover")
                    if cat == "Riparian Area" and ub==False:
                        existingRiparianFile = eflst[0]
                        totalFound +=1
                        arcpy.AddMessage("Found Riparian")
                    if cat == "Protected Area":
                        protectedAreaFile = eflst[0]
                        #totalFound +=1
                        arcpy.AddMessage("Found Protected Area")
                    if cat == "Areas of Minimum Soil Compaction":
                        areasOfMinCompacFile = eflst[0]
                        #totalFound +=1
                        arcpy.AddMessage("Found Areas of Minimum Soil Compaction")
        useExisting = False
        if existingImperviousFile !="" and proposedImperviousFile =="":
            #arcpy.AddMessage("Using Impervious from Land Cover for Parking Impervious because no Parking/Road Impervious file was found.")
            #proposedImperviousFile = existingImperviousFile
            totalFound +=1
            useExisting = True

        #if totalFound != len(cats)-2:
            #arcpy.AddMessage("Missing input file. Must have Soil, Buildings, Elevation, Parking and Roads Impervious Areas, and Impervious Areas from Land Cover.")

        #else:
        clippingBoundary = None
        if protectedAreaFile != "":
            prjBoundaryGeom =arcpy.CopyFeatures_management(self.ProjectBoundsDBPath,arcpy.Geometry())
            removeGeometry = arcpy.CopyFeatures_management(protectedAreaFile,arcpy.Geometry())
            projectWithProtectedAreaRemoved = []
            for geom in prjBoundaryGeom:
                print(geom)
                for geomT in removeGeometry:
                    geomT_sr = geomT.projectAs(geom.spatialReference)
                    geom = geom.difference(geomT_sr)
                projectWithProtectedAreaRemoved.append(geom)
            clippingBoundary = arcpy.Dissolve_management(projectWithProtectedAreaRemoved,"projectareaprotectedarea_dissolve")[0]
        else:
            clippingBoundary = self.ProjectBoundsDBPath


        uniqueName = arcpy.CreateUniqueName("criteriaAreaOutput",self.ProjectDatabase)
        criteriaAreasTable = arcpy.CreateTable_management(self.ProjectDatabase,os.path.basename(uniqueName))[0]
        arcpy.AddField_management(criteriaAreasTable,"sitingName","TEXT",field_length=150)

        arcpy.AddField_management(criteriaAreasTable,"areaAcre","DOUBLE")
        arcpy.AddField_management(criteriaAreasTable, "outDesc", "TEXT", field_length=200)
        arcpy.AddField_management(criteriaAreasTable, "outCode", "TEXT", field_length=20)
        outFields = ["sitingName", "areaAcre", "outDesc", "outCode"]
        outAreas = {}
        try:
            arcpy.AddMessage("Calculate building area values")
            buildingArea = self.calculateAcreAreaFromPolygons(buildingFile)
            outAreas["Result Directly at Source for Roof Runoff"]={"value":buildingArea}
        except:
            arcpy.AddMessage("ERROR calculating Result Directly at Source for Roof Runoff")
        try:
            #if useExisting == False:
            arcpy.AddMessage("Calculate parking area values")
            parkingArea = self.calculateAcreAreaFromPolygons(proposedImperviousFile)
            outAreas["Result Directly at Source for Parking and Road Runoff"]={"value":parkingArea}
        except:
            arcpy.AddMessage("ERROR calculating Result Directly at Source for Parking and Road Runoff")
            #with arcpy.da.InsertCursor(criteriaAreasTable,outFields) as ic:
                #ic.insertRow(["Result Directly at Source for Roof Runoff",buildingArea,"Building/roof area [for volume and placement of green roof, rainwater harvesting]","bmp_1"])
                #if useExisting == False:
                    #ic.insertRow(["Result Directly at Source for Parking and Road Runoff ",parkingArea,"Parking and road area [for volume and use of pervious pavement]","bmp_4"])





        #exclusionFC = [buildingFile,proposedImperviousFile,existingImperviousFile]
        #SLOPE CRITERIA
        try:
            arcpy.AddMessage("Creating Slope Criteria")
            slopeFivePer = self.slopeClass(demFile,"<=",5,elevCode)
            slopeFifteenPer = self.slopeClass(demFile,"<=",15,elevCode)
        except:
            arcpy.AddMessage("Missing DEM....")
        #SOIL CRITERIA
        #try:
            #arcpy.AddMessage("Creating Soil Criteria")
            #soilLayerAB = arcpy.MakeFeatureLayer_management(soilFile,"soilhydgrplayerAB",where_clause="hydgrp LIKE 'A%' or hydgrp LIKE 'B%'")[0]
            #soilLayerCD = arcpy.MakeFeatureLayer_management(soilFile,"soilhydgrplayerCD",where_clause="hydgrp LIKE 'C%' or hydgrp LIKE '%D%'")[0]
        #except:
            #arcpy.AddMessage("Missing Soil....")
        #SOIL CRITERIA
        arcpy.AddMessage("Creating Exclusion Criteria")


        if useExisting == False:
            if arcpy.Exists(protectedAreaFile):
                clip_existing = arcpy.Clip_analysis(existingImperviousFile,protectedAreaFile,"ProtectedAreaImpervious")[0]
            else:
                clip_existing = existingImperviousFile
            if buildingFile != "":
                arcpy.env.overwriteOutput = True
                allImpervious_union = arcpy.Union_analysis([buildingFile,proposedImperviousFile,clip_existing],"allImpervious_union")[0]
                allImpervious_dissolve = arcpy.Dissolve_management(allImpervious_union,'allImpervious_diss')[0]
            else:
                arcpy.env.overwriteOutput = True
                allImpervious_union = arcpy.Union_analysis([proposedImperviousFile,clip_existing],"allImpervious_union")[0]
                allImpervious_dissolve = arcpy.Dissolve_management(allImpervious_union,'allImpervious_diss')[0]
        else:
            if buildingFile != "":
                arcpy.env.overwriteOutput = True
                allImpervious_union = arcpy.Union_analysis([buildingFile,existingImperviousFile],"allImpervious_union")[0]
                allImpervious_dissolve = arcpy.Dissolve_management(allImpervious_union,'allImpervious_diss')[0]
            else:
                arcpy.env.overwriteOutput = True
                allImpervious_union = arcpy.Union_analysis([existingImperviousFile],"allImpervious_union")[0]
                allImpervious_dissolve = arcpy.Dissolve_management(allImpervious_union,'allImpervious_diss')[0]
        #Buffer Criteria
        arcpy.AddMessage("Creating Distance Criteria")
        try:
            buildingLayer = arcpy.MakeFeatureLayer_management(buildingFile,"buildingslayer")[0]
            buildingBetween10and50 = self.bufferBetween(buildingLayer,10,50,"building10to50")
            buildingLT50 = self.bufferFull(buildingLayer,50,"buildingto50")
        except:
            pass

        try:
            arcpy.AddMessage("parking buffers")
            if useExisting == False:
                arcpy.AddMessage("use existing false")
                parkingLayer = arcpy.MakeFeatureLayer_management(proposedImperviousFile,"parkingLayer")[0]
                arcpy.AddMessage("buffer parking 1 to 50")
                parkingBetween1and50 = self.bufferBetween(parkingLayer,1,50,"parking1to50")
                arcpy.AddMessage("buffer parking 1 to 10")
                parkingBetween1and10 = self.bufferBetween(parkingLayer,1,10,"parking1to10")
                arcpy.AddMessage("buffer parking to 50")
                parkingLT50 = self.bufferFull(parkingLayer,50,"parkingto50")
        except:
            arcpy.AddMessage("Unexpected error:%s"%sys.exc_info()[0])
            arcpy.AddMessage("Imervious area and riparian area buffer ERROR")

        try:
            arcpy.AddMessage("all impervious")
            allImperviousGT50 = self.inverseBuffer(allImpervious_dissolve,50,"FiftyFtFromAllImpervious",self.ProjectBoundsDBPath)


            arcpy.AddMessage("riparian buffer")
            riparianGT1 = self.inverseBuffer(existingRiparianFile,1,"OneFTFromRiparian",self.ProjectBoundsDBPath)
            riparianBetween1and20 = self.bufferBetween(existingRiparianFile,1,20,"riparian1to20")

        except:
            arcpy.AddMessage("Unexpected error:%s"%sys.exc_info()[0])
            arcpy.AddMessage("Imervious area and riparian area buffer ERROR")

        #First One
        try:
            result = self.interDissExclude([buildingBetween10and50,slopeFivePer],[allImpervious_dissolve],"c_rr_raingardenbioswale",clippingBoundary)
            self.resultsFileOrg([[result,"Result Rain Gardens and Bioswales Roof Runoff"]])
            area = self.calculateAcreAreaFromPolygons(result)
            outAreas["Result Rain Gardens and Bioswales Roof Runoff"]={"value":area}
            resultsList["Result Rain Gardens and Bioswales Roof Runoff"]=[result,"bmp_2"]
            #with arcpy.da.InsertCursor(criteriaAreasTable,outFields) as ic:
                #ic.insertRow(["Result Rain Gardens and Bioswales Roof Runoff",area,r"< 50 ft and > 10 ft from buildings; A&B Soils; <5% slope, excluding all impervious areas [for placement of rain gardens and bioswales]","bmp_2"])
        except:
            arcpy.AddMessage("Unexpected error:%s"%sys.exc_info()[0])
            arcpy.AddMessage("ERROR: Result Rain Gardens and Bioswales Roof Runoff")


        try:
            arcpy.AddMessage("Result Trees At Source Roof Runoff")
            result = self.interDissExclude([buildingBetween10and50],[allImpervious_dissolve],"c_rr_trees",clippingBoundary)
            self.resultsFileOrg([[result,"Result Trees At Source Roof Runoff"]])
            area = self.calculateAcreAreaFromPolygons(result)
            outAreas["Result Trees At Source Roof Runoff"]={"value":area}
            resultsList["Result Trees At Source Roof Runoff"]=[result,"bmp_3"]
            #with arcpy.da.InsertCursor(criteriaAreasTable,outFields) as ic:
                #ic.insertRow(["Result Trees At Source Roof Runoff",area,r"< 50 ft and >10ft from buildings; ALL Soils; ALL slope, excluding all impervious areas [for placement of trees to manage roof runoff]","bmp_3"])
        except:
            arcpy.AddMessage("Unexpected error:%s"%sys.exc_info()[0])
            arcpy.AddMessage("ERROR: Result Trees At Source Roof Runoff")

        if useExisting == False:
            try:
                arcpy.AddMessage("Result Rain Gardens and Bioswales at Parking and Road Runoff")
                result = self.interDissExclude([parkingBetween1and50,slopeFivePer],[allImpervious_dissolve,buildingLT50],"c_pr_raingardenbioswale",clippingBoundary)
                self.resultsFileOrg([[result,"Result Rain Gardens and Bioswales at Parking and Road Runoff"]])
                area = self.calculateAcreAreaFromPolygons(result)
                outAreas["Result Rain Gardens and Bioswales at Parking and Road Runoff"]={"value":area}
                resultsList["Result Rain Gardens and Bioswales at Parking and Road Runoff"]=[result,"bmp_5"]
                #with arcpy.da.InsertCursor(criteriaAreasTable,outFields) as ic:
                    #ic.insertRow(["Result Rain Gardens and Bioswales at Parking and Road Runoff",area,"< 50 ft and > 1 ft from roads and parking areas; A&B Soils; <5% slope; excluding ALL impervious areas and ALL areas <50 ft of a building [for placement of rain gardens and bioswales]","bmp_5"])
            except:
                arcpy.AddMessage("Unexpected error:%s"%sys.exc_info()[0])
                arcpy.AddMessage("ERROR: Result Rain Gardens and Bioswales at Parking and Road Runoff")

            try:
                arcpy.AddMessage("Result Trees at Source for Parking and Road Runoff")
                result = self.interDissExclude([parkingLT50,slopeFivePer],[allImpervious_dissolve,buildingLT50],"c_pr_trees",clippingBoundary)
                self.resultsFileOrg([[result,"Result Trees at Source for Parking and Road Runoff"]])
                area = self.calculateAcreAreaFromPolygons(result)
                outAreas["Result Trees at Source for Parking and Road Runoff"]={"value":area}
                resultsList["Result Trees at Source for Parking and Road Runoff"]=[result,"bmp_7"]
                #with arcpy.da.InsertCursor(criteriaAreasTable,outFields) as ic:
                    #ic.insertRow(["Result Trees at Source for Parking and Road Runoff",area,r"< 50 ft from roads and parking areas; ALL Soils; <5% slope; excluding ALL impervious areas and ALL areas <50 ft of a building [for placement of Trees]","bmp_7"])
            except:
                arcpy.AddMessage("Unexpected error:%s"%sys.exc_info()[0])
                arcpy.AddMessage("ERROR: Result Trees at Source for Parking and Road Runoff")

            try:
                arcpy.AddMessage("Result Infiltration Trenches at Parking and Road Runoff")
                result = self.interDissExclude([parkingBetween1and50,slopeFivePer],[allImpervious_dissolve,buildingLT50],"c_pr_trench",clippingBoundary)
                self.resultsFileOrg([[result,"Result Infiltration Trenches at Parking and Road Runoff"]])
                area = self.calculateAcreAreaFromPolygons(result)
                outAreas["Result Infiltration Trenches at Parking and Road Runoff"]={"value":area}
                resultsList["Result Infiltration Trenches at Parking and Road Runoff"]=[result,"bmp_6"]
                #with arcpy.da.InsertCursor(criteriaAreasTable,outFields) as ic:
                    #ic.insertRow(["Result Infiltration Trenches at Parking and Road Runoff",area,r"< 50 ft and > 1 ft from roads and parking areas; A&B Soils; <5%; excluding ALL impervious areas and ALL areas <50 ft of a building [for placement of Infiltration Trenches]","bmp_6"])
            except:
                arcpy.AddMessage("Unexpected error:%s"%sys.exc_info()[0])
                arcpy.AddMessage("ERROR: Result Infiltration Trenches at Parking and Road Runoff")

        try:
            arcpy.AddMessage("Result Infiltration Basin for All Pervious")
            result = self.interDissExclude([allImperviousGT50,riparianGT1,slopeFivePer],[],"c_ai_infilbasin",clippingBoundary)
            self.resultsFileOrg([[result,"Result Infiltration Basin for All Pervious"]])
            area = self.calculateAcreAreaFromPolygons(result)
            outAreas["Result Infiltration Basin for All Pervious"]={"value":area}
            resultsList["Result Infiltration Basin for All Pervious"]=[result,"bmp_8"]
            #with arcpy.da.InsertCursor(criteriaAreasTable,outFields) as ic:
                #ic.insertRow(["Result Infiltration Basin for All Impervious",area,r">50 ft from ALL impervious areas and > 1 ft from riparian areas; A&B Soils; <5% slope [infiltration basin]","bmp_8"])
        except:
            arcpy.AddMessage("Unexpected error:%s"%sys.exc_info()[0])
            arcpy.AddMessage("ERROR: Result Infiltration Basin for All Pervious")

        try:
            arcpy.AddMessage("Result Bioswales for All Pervious")
            result = self.interDissExclude([allImperviousGT50,riparianGT1,slopeFivePer],[],"c_ai_bioswale",clippingBoundary)
            self.resultsFileOrg([[result,"Result Bioswales for All Pervious"]])
            area = self.calculateAcreAreaFromPolygons(result)
            outAreas["Result Bioswales for All Pervious"]={"value":area}
            resultsList["Result Bioswales for All Pervious"]=[result,"bmp_9"]
            #with arcpy.da.InsertCursor(criteriaAreasTable,outFields) as ic:
                #ic.insertRow(["Result Bioswales for All Impervious",area,">50 ft from ALL impervious areas and > 1 ft from riparian areas; A&B Soils; 5% slope [Bioswale]","bmp_9"])
        except:
            arcpy.AddMessage("Unexpected error:%s"%sys.exc_info()[0])
            arcpy.AddMessage("ERROR: Result Bioswales for All Pervious")

        try:
            arcpy.AddMessage("Result Constructed Wetland for All Pervious")
            result = self.interDissExclude([allImperviousGT50,riparianGT1,slopeFifteenPer],[],"c_ai_constwetlands",clippingBoundary)
            self.resultsFileOrg([[result,"Result Constructed Wetland for All Pervious"]])
            area = self.calculateAcreAreaFromPolygons(result)
            outAreas["Result Constructed Wetland for All Pervious"]={"value":area}
            resultsList["Result Constructed Wetland for All Pervious"]=[result,"bmp_10"]
            #with arcpy.da.InsertCursor(criteriaAreasTable,outFields) as ic:
                #ic.insertRow(["Result Constructed Wetland for All Impervious",area,r">50 ft from ALL impervious areas and > 1 ft from riparian areas; C&D including AD, BD, and CD Soils; <15% slope [for constructed wetland]","bmp_10"])
        except:
            arcpy.AddMessage("Unexpected error:%s"%sys.exc_info()[0])
            arcpy.AddMessage("ERROR: Result Constructed Wetland for All Pervious")

        try:
            arcpy.AddMessage("Result Wet Ponds and Dry Ponds for All Pervious")
            result = self.interDissExclude([allImperviousGT50,riparianGT1,slopeFifteenPer],[],"c_ai_wetdrypond",clippingBoundary)
            self.resultsFileOrg([[result,"Result Wet Ponds and Dry Ponds for All Pervious"]])
            area = self.calculateAcreAreaFromPolygons(result)
            outAreas["Result Wet Ponds and Dry Ponds for All Pervious"]={"value":area}
            resultsList["Result Wet Ponds and Dry Ponds for All Pervious"]=[result,"bmp_11"]
            #with arcpy.da.InsertCursor(criteriaAreasTable,outFields) as ic:
                #ic.insertRow(["Result Wet Ponds and Dry Ponds for All Impervious",area,r">50 ft from ALL impervious areas and > 1 ft from riparian areas; ALL Soils; <15% slope [for wet pond, dry pond]","bmp_11"])
        except:
            arcpy.AddMessage("Unexpected error:%s"%sys.exc_info()[0])
            arcpy.AddMessage("ERROR: Result Wet Ponds and Dry Ponds for All Pervious")

        try:
            arcpy.AddMessage("Result Trees for All Pervious")
            result = self.interDissExclude([allImperviousGT50,slopeFivePer],[],"c_ai_trees",clippingBoundary)
            self.resultsFileOrg([[result,"Result All Pervious Trees"]])
            area = self.calculateAcreAreaFromPolygons(result)
            outAreas["Result Trees for All Pervious"]={"value":area}
            resultsList["Result Trees for All Pervious"]=[result,"bmp_12"]
            #with arcpy.da.InsertCursor(criteriaAreasTable,outFields) as ic:
                #ic.insertRow(["Result Trees for All Impervious",area,r">50 ft from ALL impervious areas; All Soils; <5% slope [for Trees]","bmp_12"])
        except:
            arcpy.AddMessage("Unexpected error:%s"%sys.exc_info()[0])
            arcpy.AddMessage("ERROR: Result Trees for All Pervious")

        #Originally calculating overlap between the different results.

        #with arcpy.da.SearchCursor(self.ProjectDatabase+"\\FileTypesResults",["FileName","LayerOrder"],where_clause="Category = 'Criteria Result'",sql_clause=(None, 'ORDER BY LayerOrder')) as sc:
            #key_order = []
            #for row in sc:
                #key_order.append(row[0])
        #for i,k in enumerate(key_order):
            #if k in resultsList.keys():
                #file1 = resultsList[k][0]
                #code1 = resultsList[k][1]
                #for j in range(i,len(key_order)):
                    #k2 = key_order[j]
                    #if k2 in resultsList.keys():
                        #if k!=k2:
                            #file2 = resultsList[k2][0]
                            #code2 = resultsList[k2][1]
                            #arcpy.AddMessage(file1)
                            #arcpy.AddMessage(file2)
                            #overlapArea = self.overlapArea(file1,file2)
                            #with arcpy.da.InsertCursor(criteriaAreasTable,outFields) as ic:
                                #ic.insertRow(["%s overlaps with %s"%(code1,code2),overlapArea,"%s overlaps with %s"%(k,k2),"%s_%s"%(code1,code2)])


        #No Longer Writing directly to excel file.

        #fileName = self.ProjectWorkbook#self.ProjectFolder + "\\" + self.ProjectName + ".xlsx"
        #wb = load_workbook(fileName)
        #ws = wb['DataDictionary']
        #rowIndx = ws.max_row+1

        #with arcpy.da.SearchCursor(criteriaAreasTable,outFields) as sc:
            #for row in sc:
                #ws.cell(row = rowIndx, column = 1).value = row[0]
                #ws.cell(row = rowIndx, column = 2).value = row[1]
                #ws.cell(row = rowIndx, column = 3).value = "Acres"
                #ws.cell(row = rowIndx, column = 4).value = row[2]
                #ws.cell(row = rowIndx, column = 5).value = row[3]
                #rowIndx +=1
        #wb.save(fileName)
        self.writeValuesToAreasTable(outAreas)


    def calculateAcreAreaFromPolygons(self,fc,units="ACRES"):
        currentArea = 0.0
        with arcpy.da.SearchCursor(fc,["SHAPE@"]) as sc:
            for row in sc:
                poly = row[0]
                currentArea += poly.getArea("PLANAR",units)
        return currentArea

    def inverseBuffer(self,ef,distance,outName,boundaryExtent):
        polygon = self.buildPolygonFromFCExtent(boundaryExtent)
        arcpy.AddMessage("Area %s"%(polygon.area))
        resultGeom = arcpy.Buffer_analysis(ef,arcpy.Geometry(),"%s FEET"%(distance),dissolve_option="ALL")[0]
        polygon = polygon.projectAs(resultGeom.spatialReference)
        arcpy.RepairGeometry_management(resultGeom)
        arcpy.RepairGeometry_management(polygon)
        cutout = polygon.difference(resultGeom)
        arcpy.env.overwriteOutput=True
        arcpy.env.workspace = self.ProjectDatabase
        extentPoly = arcpy.CopyFeatures_management(polygon,"ProjectBoundsExtentPoly")[0]
        res = arcpy.CopyFeatures_management(cutout,outName)[0]
        arcpy.RepairGeometry_management(res)
        return res

    def bufferFull(self,ef,distance,outName):
        """Distance is in Feet"""
        resultGeom = arcpy.Buffer_analysis(ef,outName,"%s FEET"%(distance),dissolve_option="ALL")[0]
        arcpy.RepairGeometry_management(resultGeom)
        return resultGeom

    def bufferBetween(self,ef,startDistance,endDistance,outName):
        """Distances is in Feet"""
        arcpy.env.workspace = self.ProjectDatabase
        bufferStart = arcpy.Buffer_analysis(ef,"buffer_%sft"%startDistance,"%s FEET"%(startDistance),"FULL","ROUND","NONE")[0]
        arcpy.RepairGeometry_management(bufferStart)
        diff = endDistance - startDistance
        bufferOutside = arcpy.Buffer_analysis(bufferStart,outName,"%s FEET"%(diff),"OUTSIDE_ONLY","ROUND","ALL")[0]
        arcpy.RepairGeometry_management(bufferOutside)
        return bufferOutside

    def slopeClass(self,dem,symbol,percentSlope,elevCode):
        arcpy.env.workspace = self.ProjectDatabase
        sr = arcpy.Describe(dem).spatialReference

        zF = sharedTools.zfactorConvertZtoLinear(sr.linearUnitCode,elevCode)
        rast = arcpy.Raster(dem)
        symText = ""
        if "<" in symbol:
            symText = "lt"
        elif ">" in symbol:
            symText = "gt"
        else:
            symText = "eq"
        try:
            rast_slope = arcpy.sa.Slope(rast,"PERCENT_RISE",zF)
            rast_slope.save("slope_rast")
            res = arcpy.sa.Int(arcpy.sa.Con(rast_slope,1,where_clause="VALUE%s%s"%(symbol,percentSlope)))
            res_slope = arcpy.sa.ExtractByAttributes(res,"VALUE = 1")
            res_slope.save("r_slope_per%s%s"%(symText,percentSlope))
            arcpy.Delete_management(res)
            res = arcpy.RasterToPolygon_conversion(res_slope,"slope%s%s_frmrast"%(symText,percentSlope),"NO_SIMPLIFY","VALUE")[0]
            arcpy.RepairGeometry_management(res)
            self._deleteList.append(res)
            res = arcpy.Clip_analysis(res,self.ProjectBoundsDBPath,"slope%s%s_clip"%(symText,percentSlope))[0]
            return res
        except:
            return None

    def interDissExclude(self,files,exclusionAreas,outName,clipBoundary=None):
        """Both are lists of feature Classes"""
        arcpy.env.workspace = self.ProjectDatabase
        arcpy.env.overwriteOutput=True
        for f in files:
            arcpy.RepairGeometry_management(f)
        if len(files)>1:
            intersectResults = arcpy.Intersect_analysis(files,"Tempint%s"%outName)[0]
        if len(files) == 1:
            intersectResults = files[0]
        if len(files) == 0:
            return None
        #print intersectResults
        self._deleteList.append(intersectResults)
        dissolveResults = arcpy.Dissolve_management(intersectResults,arcpy.Geometry())
        #print dissolveResults
        finalGeom = []
        for geom_1 in dissolveResults:
            for fc in exclusionAreas:
                with arcpy.da.SearchCursor(fc,["SHAPE@"]) as sc:
                    for row in sc:
                        geom_2 = row[0]
                        geom_1 = geom_1.difference(geom_2) #cut out anything that overlaps and return it to geom_1
            finalGeom.append(geom_1)
        #print len(finalGeom)
        if clipBoundary:
            outFCT = arcpy.CreateFeatureclass_management(self.ProjectDatabase,outName+"_temp","POLYGON",spatial_reference=self.getSpatialReference())[0]
            with arcpy.da.InsertCursor(outFCT,["SHAPE@"]) as ic:
                for geom in finalGeom:
                    ic.insertRow([geom])
            outFC = arcpy.Clip_analysis(outFCT,clipBoundary,outName)[0]
        else:
            outFC = arcpy.CreateFeatureclass_management(self.ProjectDatabase,outName,"POLYGON",spatial_reference=self.getSpatialReference())[0]
            with arcpy.da.InsertCursor(outFC,["SHAPE@"]) as ic:
                for geom in finalGeom:
                    ic.insertRow([geom])

        arcpy.RepairGeometry_management(outFC)
        return outFC
    def overlapArea(self,file1,file2):
        """Both are lists of feature Classes"""
        arcpy.env.workspace = self.ProjectDatabase
        arcpy.env.overwriteOutput=True
        arcpy.RepairGeometry_management(file1)
        arcpy.RepairGeometry_management(file2)
        with arcpy.da.SearchCursor(file1,["SHAPE@"]) as sc:
            for row in sc:
                geom_1 = row[0]
                break
        with arcpy.da.SearchCursor(file2,["SHAPE@"]) as sc:
            for row in sc:
                geom_2 = row[0]
                break

        try:
            geom_2 = geom_2.projectAs(geom_1.spatialReference)
            if geom_1.disjoint(geom_2)==True:
                arcpy.AddMessage("disjoint is true")
                return 0.0
            else:
                geom_inter = geom_1.intersect(geom_2,4)
                arcpy.AddMessage("intersected")
                area = geom_inter.getArea("PLANAR","ACRES")
                arcpy.AddMessage(area)
                return area
        except:
            arcpy.AddMessage("Still not working")
            return 0.0



    def treesWithinImperviousPervious(self):
        arcpy.AddMessage("==============Pervious/Impervious Soil Types==============")
        arcpy.env.workspace = self.ProjectDatabase
        arcpy.env.overwriteOutput=True
        settings = self.getSettingsDictionary()
        elevCode = sharedTools.ELEVATION_UNITS[settings[self.ELEVATION_UNIT_KEY]]
        arcpy.AddMessage("Checking for input files...")
        cats = ["Soil","Impervious From Land Cover","Pervious From Land Cover","Parking Impervious","Buildings","Protected Area","Trees From Land Cover"]
        soilFile = ""
        perviousFile = ""
        proposedImperviousFile = ""
        existingImperviousFile = ""
        buildingsFile = ""
        protectedAreaFile = ""
        treesFile = ""
        totalFound = 0
        for cat in cats:
            fileTypes = self.getFileTypesFromCat(cat)
            for ft,geo,ub in fileTypes:
                eflst = self.getFiles(ft)
                if len(eflst)==1:
                    if cat == "Soil":
                        soilFile = eflst[0]
                        totalFound +=1
                        arcpy.AddMessage("Found Soil")
                    if cat == "Buildings":
                        buildingFile = eflst[0]
                        totalFound +=1
                        arcpy.AddMessage("Found Buildings")
                    if cat == "Pervious From Land Cover":
                        perviousFile = eflst[0]
                        totalFound +=1
                        arcpy.AddMessage("Found Pervious From Land Cover")
                    if cat == "Parking Impervious":
                        proposedImperviousFile = eflst[0]
                        totalFound +=1
                        arcpy.AddMessage("Found Parking Impervious")
                    if cat == "Impervious From Land Cover":
                        existingImperviousFile = eflst[0]
                        totalFound +=1
                        arcpy.AddMessage("Found Impervious from Land Cover")
                    if cat == "Protected Area":
                        protectedAreaFile = eflst[0]
                        #totalFound +=1
                        arcpy.AddMessage("Found Protected Area")
                    if cat == "Trees From Land Cover":
                        treesFile = eflst[0]
                        totalFound +=1
                        arcpy.AddMessage("Found Trees From Land Cover")

        useExisting = False
        if existingImperviousFile !=None and proposedImperviousFile ==None:
            #arcpy.AddMessage("Using Impervious from Land Cover for Parking Impervious because no Parking/Road Impervious file was found.")
            #proposedImperviousFile = existingImperviousFile
            totalFound +=1
            useExisting = True

        if totalFound != len(cats)-1:
            arcpy.AddMessage("Missing input file. Must have Soil, Buildings, Pervious from Land Cover,  and Impervious Areas from Land Cover.")
        else:
            uniqueName = arcpy.CreateUniqueName("soilPerviousAreaOutput",self.ProjectDatabase)
            arcpy.AddMessage(uniqueName)
            criteriaAreasTable = arcpy.CreateTable_management(self.ProjectDatabase,os.path.basename(uniqueName))[0]
            arcpy.AddField_management(criteriaAreasTable,"typeName","TEXT",field_length=150)
            arcpy.AddField_management(criteriaAreasTable,"areaAcre","DOUBLE")

            arcpy.AddMessage("Creating Trees Layer")
            treesLayerAll = arcpy.MakeFeatureLayer_management(soilFile,"treeslayerAll")[0]

            #SOIL CRITERIA
            arcpy.AddMessage("Creating Soil Layer")
            soilLayerAll = arcpy.MakeFeatureLayer_management(soilFile,"soilhydgrplayerAll")[0]
            #SOIL CRITERIA
            arcpy.AddMessage("Creating Impervious")
            arcpy.AddMessage(existingImperviousFile)
            arcpy.AddMessage(protectedAreaFile)

            if arcpy.Exists(protectedAreaFile):
                clip_existing = arcpy.Clip_analysis(existingImperviousFile,protectedAreaFile,"ProtectedAreaImpervious2")[0]
            else:
                clip_existing = existingImperviousFile

            if useExisting == True:

                allImpervious_union = arcpy.Union_analysis([buildingFile,clip_existing],"allImpervious_union2")[0]
                allImpervious_geometry = arcpy.Dissolve_management(allImpervious_union,arcpy.Geometry())[0]
            else:
                #allImpervious_union = arcpy.Union_analysis([buildingFile,existingImperviousFile],"allImpervious_union")[0]
                #allImpervious_dissolve = arcpy.Dissolve_management(allImpervious_union,'allImpervious_diss')[0]
                allImpervious_union = arcpy.Union_analysis([buildingFile,proposedImperviousFile,clip_existing],"allImpervious_union2")[0]
                allImpervious_geometry = arcpy.Dissolve_management(allImpervious_union,arcpy.Geometry())[0]


            # arcpy.AddMessage(buildingFile)
            # arcpy.AddMessage(clip_existingImpervious)
            # allImpervious_union = arcpy.Union_analysis([buildingFile,proposedImperviousFile,clip_existingImpervious],"allImpervious_union")[0]
            # allImpervious_geometry = arcpy.Dissolve_management(allImpervious_union,arcpy.Geometry())[0]

            perviousGeoms = []
            with arcpy.da.SearchCursor(self.ProjectBoundsDBPath,["SHAPE@"]) as sc:
                for row in sc:
                    poly = row[0]
                    allImpervious_geometry_sr = allImpervious_geometry.projectAs(poly.spatialReference)
                    perviousGeoms.append(poly.difference(allImpervious_geometry_sr))

            if len(perviousGeoms) >0:
                perviousDissolve = arcpy.Dissolve_management(perviousGeoms,"allPervious_dissolve")[0]
                allImpervious_dissolve = arcpy.CopyFeatures_management(allImpervious_geometry,"allImpervious_dissolve")[0]
                perviousTreesSoil = arcpy.Intersect_analysis([perviousDissolve,treesLayerAll,soilLayerAll],"pervious_trees_intersect")[0]
                perviousTreesSoil = arcpy.Dissolve_management(perviousTreesSoil,"pervious_trees_soil","HYDGRP")[0]
                output = []
                arcpy.AddMessage("Write pervious trees soil....")
                with arcpy.da.SearchCursor(perviousTreesSoil,["SHAPE@","HYDGRP"]) as sc:
                    for row in sc:
                        poly = row[0]
                        output.append(["Pervious (forest) %s"%row[1],poly.getArea("PLANAR","ACRES")])

                imperviousSoil = arcpy.Intersect_analysis([allImpervious_dissolve,soilLayerAll],"impervious_intersect")[0]
                imperviousSoil = arcpy.Dissolve_management(imperviousSoil,"impervious_soil","HYDGRP")[0]
                arcpy.AddMessage("Write impervious soil....")
                with arcpy.da.SearchCursor(imperviousSoil,["SHAPE@","HYDGRP"]) as sc:
                    for row in sc:
                        poly = row[0]
                        output.append(["Impervious %s"%row[1],poly.getArea("PLANAR","ACRES")])

                perviousGeomsNoTrees = []
                treesGeometry = arcpy.CopyFeatures_management(treesFile,arcpy.Geometry())
                treesFileSR = arcpy.Describe(treesFile).spatialReference
                for perviousGeom in perviousGeoms:
                    for tree in treesGeometry:
                        tree_sr = tree.projectAs(perviousGeom.spatialReference)
                        perviousGeom = perviousGeom.difference(tree_sr)
                    perviousGeomsNoTrees.append(perviousGeom)

                perviousNoTreesDissolve = arcpy.Dissolve_management(perviousGeomsNoTrees,"allPerviousNoTrees_dissolve")[0]
                perviousNoTreesSoil = arcpy.Intersect_analysis([perviousNoTreesDissolve,soilLayerAll],"pervious_notrees_intersect")[0]
                perviousNoTreesSoil = arcpy.Dissolve_management(perviousNoTreesSoil,"pervious_notrees_soil","HYDGRP")[0]

                arcpy.AddMessage("Write pervious no trees soil....")
                with arcpy.da.SearchCursor(perviousNoTreesSoil,["SHAPE@","HYDGRP"]) as sc:
                    for row in sc:
                        poly = row[0]
                        output.append(["Pervious (no forest) %s"%row[1],poly.getArea("PLANAR","ACRES")])

                fileName = self.ProjectWorkbook
                #wb = load_workbook(fileName)
                #ws = wb['DataDictionary']
                #rowIndx = ws.max_row+1

                #with arcpy.da.InsertCursor(criteriaAreasTable,["typeName","areaAcre"]) as ic:
                    #for o in output:
                        #ic.insertRow(o)
                        #ws.cell(row = rowIndx, column = 1).value = o[0]
                        #ws.cell(row = rowIndx, column = 2).value = o[1]
                        #ws.cell(row = rowIndx, column = 3).value = "Acres"
                        #rowIndx +=1
                #wb.save(fileName)
    def imperviousPerviousAreas(self):
        arcpy.AddMessage("==============Identify Impervious Pervious Areas==============")
        arcpy.env.workspace = self.ProjectDatabase
        arcpy.env.overwriteOutput=True
        settings = self.getSettingsDictionary()
        elevCode = sharedTools.ELEVATION_UNITS[settings[self.ELEVATION_UNIT_KEY]]
        arcpy.AddMessage("Checking for input files...")
        cats = ["Buildings","Parking Impervious","Impervious From Land Cover","Trees From Land Cover", "Pervious From Land Cover","Protected Area"]

        perviousFile = ""
        landcoverImperviousFile = ""
        polygonImperviousFile = ""
        buildingFile = ""
        treesFile = ""
        protectedAreaFile = ""
        for cat in cats:
            fileTypes = self.getFileTypesFromCat(cat)
            for ft,geo,ub in fileTypes:
                eflst = self.getFiles(ft)
                if len(eflst)==1:
                    if cat == "Buildings":
                        buildingFile = eflst[0]
                        arcpy.AddMessage("Found Buildings")
                    if cat == "Parking Impervious":
                        proposedImperviousFile = eflst[0]
                        arcpy.AddMessage("Found Parking Impervious")
                    if cat == "Impervious From Land Cover":
                        existingImperviousFile = eflst[0]
                        arcpy.AddMessage("Found Impervious from Land Cover")
                    if cat == "Trees From Land Cover":
                        treesFile = eflst[0]
                        arcpy.AddMessage("Found Trees from Land Cover")
                    if cat == "Pervious From Land Cover":
                        perviousFile = eflst[0]
                        arcpy.AddMessage("Found Pervious from Land Cover")
                    if cat == "Protected Area":
                        protectedAreaFile = eflst[0]
                        #totalFound +=1
                        arcpy.AddMessage("Found Protected Area")
        areas = {"Total Pervious From Landcover":{"value":0},
        "Impervious (Landcover)":{"value":0},
        "Total Impervious From Landcover, Building Footprints, and Parking and Roads":{"value":0},
        "Impervious (Parking, Road if Available)":{"value":0},
        "Impervious (Building Footprints if Available)":{"value":0},
        "Pervious (Trees)":{"value":0},
        "Pervious (No Trees)":{"value":0},
        "Impervious Landcover Minus Buildings, Roads and parking":{"value":0},
        "Total Pervious From Landcover Minus Building Footprints and Parking and Roads":{"value":0},
        "Impervious Landcover Minus Buildings":{"value":0},
        }

        clippingBoundary = None
        if protectedAreaFile != "":
            prjBoundaryGeom =arcpy.CopyFeatures_management(self.ProjectBoundsDBPath,arcpy.Geometry())
            removeGeometry = arcpy.CopyFeatures_management(protectedAreaFile,arcpy.Geometry())
            projectWithProtectedAreaRemoved = []
            for geom in prjBoundaryGeom:
                print(geom)
                for geomT in removeGeometry:
                    geomT_sr = geomT.projectAs(geom.spatialReference)
                    geom = geom.difference(geomT_sr)
                projectWithProtectedAreaRemoved.append(geom)
            clippingBoundary = arcpy.Dissolve_management(projectWithProtectedAreaRemoved,"projectareaprotectedarea_dissolve")[0]
        else:
            clippingBoundary = self.ProjectBoundsDBPath
        print(clippingBoundary)

        imperviousLandCoverLayer = None
        buildingLayer = None
        parkingImperviousLayer = None
        perviousLayer = None
        if perviousFile !="":
            areas["Total Pervious From Landcover"]["value"]=self.calculateAcreAreaFromPolygons(perviousFile)
            perviousLayer = arcpy.MakeFeatureLayer_management(perviousFile,"perviousLayerCalcAll")[0]

        if existingImperviousFile !="":
            areas["Impervious (Landcover)"]["value"]=self.calculateAcreAreaFromPolygons(existingImperviousFile)
            imperviousLandCoverLayer = arcpy.MakeFeatureLayer_management(existingImperviousFile,"imperviousLandCoverLayerAll")[0]

        if buildingFile !="":
            areas["Impervious (Building Footprints if Available)"]["value"]=self.calculateAcreAreaFromPolygons(buildingFile)
            buildingLayer = arcpy.MakeFeatureLayer_management(buildingFile,"buildingLayerAll")[0]

        if proposedImperviousFile !="":
            areas["Impervious (Parking, Road if Available)"]["value"]=self.calculateAcreAreaFromPolygons(proposedImperviousFile)
            parkingImperviousLayer = arcpy.MakeFeatureLayer_management(proposedImperviousFile,"parkingImperviousLayerAll")[0]

        if treesFile != "":
            treesLayerAll = arcpy.MakeFeatureLayer_management(treesFile,"treeslayerAll")[0]

        perviousGeomsNoImperv_clip = None


        #"Total Impervious From Landcover, Building Footprints, and Parking and Roads"
        if imperviousLandCoverLayer:
            lyrsToCombine = [parkingImperviousLayer,buildingLayer]
            for lyr in lyrsToCombine:
                if lyr == None:
                    lyrsToCombine.remove(lyr)
            if len(lyrsToCombine) > 0:

                arcpy.env.workspace = self.ProjectDatabase
                arcpy.env.overwriteOutput = True
                allLyrs = [imperviousLandCoverLayer]+lyrsToCombine
                print allLyrs
                TotalImpervious_union = arcpy.Union_analysis(allLyrs,"TotalImpervious_union")[0]
                print(TotalImpervious_union)
                TotalImpervious_dissolve = arcpy.Dissolve_management(TotalImpervious_union,'TotalImpervious_diss')[0]
                print(TotalImpervious_dissolve)
                TotalImpervious_clip = arcpy.Clip_analysis(TotalImpervious_dissolve,clippingBoundary,"TotalImpervious_clip")[0]
                print(TotalImpervious_clip)
                areas["Total Impervious From Landcover, Building Footprints, and Parking and Roads"]["value"]=self.calculateAcreAreaFromPolygons(TotalImpervious_clip)
                print(areas)
                removeGeometry = []


                imperviousLandCoverRemove_clip=None
                if buildingLayer:
                    removeGeometry = arcpy.CopyFeatures_management(buildingLayer,arcpy.Geometry())
                    imperviousGeoms = arcpy.CopyFeatures_management(imperviousLandCoverLayer,arcpy.Geometry())
                    imperviousGeomsRemove = []
                    print(len(imperviousGeoms))
                    print(len(removeGeometry))
                    for imperviousGeom in imperviousGeoms:
                        for geomT in removeGeometry:
                            geomT_sr = geomT.projectAs(imperviousGeom.spatialReference)
                            imperviousGeom = imperviousGeom.difference(geomT_sr)
                        imperviousGeomsRemove.append(imperviousGeom)

                    imperviousLandCoverRemoveDissolve = arcpy.Dissolve_management(imperviousGeomsRemove,"imperviousLandCoverRemoveBld_dissolve")[0]
                    imperviousLandCoverRemove_clip = arcpy.Clip_analysis(imperviousLandCoverRemoveDissolve,clippingBoundary,"imperviousLandCoverRemoveBld_clip")[0]
                    areas["Impervious Landcover Minus Buildings"]["value"]=self.calculateAcreAreaFromPolygons(imperviousLandCoverRemove_clip)
                    print(areas)

                if parkingImperviousLayer:
                    if imperviousLandCoverRemove_clip:
                        imperviousGeoms = arcpy.CopyFeatures_management(imperviousLandCoverRemove_clip,arcpy.Geometry())
                    else:
                        imperviousGeoms = arcpy.CopyFeatures_management(imperviousLandCoverLayer,arcpy.Geometry())
                    imperviousGeomsRemove = []
                    print(len(imperviousGeoms))
                    print(len(removeGeometry))
                    for imperviousGeom in imperviousGeoms:
                        for geomT in removeGeometry:
                            geomT_sr = geomT.projectAs(imperviousGeom.spatialReference)
                            imperviousGeom = imperviousGeom.difference(geomT_sr)
                        imperviousGeomsRemove.append(imperviousGeom)
                        print(len(imperviousGeomsRemove))
                    print("dissolving impervious areas")
                    imperviousLandCoverRemoveDissolve = arcpy.Dissolve_management(imperviousGeomsRemove,"imperviousLandCoverRemove_dissolve")[0]
                    print("clipping impervious areas")
                    imperviousLandCoverRemove_clip = arcpy.Clip_analysis(imperviousLandCoverRemoveDissolve,clippingBoundary,"imperviousLandCoverRemove_clip")[0]
                    areas["Impervious Landcover Minus Buildings, Roads and parking"]["value"]=self.calculateAcreAreaFromPolygons(imperviousLandCoverRemove_clip)
                    print(areas)

                perviousGeoms = arcpy.CopyFeatures_management(perviousFile,arcpy.Geometry())
                perviousGeomsNoImperv = []
                for perviousGeom in perviousGeoms:
                    for geomT in removeGeometry:
                        geomT_sr = geomT.projectAs(imperviousGeom.spatialReference)
                        perviousGeom = perviousGeom.difference(geomT_sr)
                    perviousGeomsNoImperv.append(perviousGeom)

                perviousGeomsNoImpervDissolve = arcpy.Dissolve_management(perviousGeomsNoImperv,"perviousGeomsNoImperv_dissolve")[0]
                perviousGeomsNoImperv_clip = arcpy.Clip_analysis(perviousGeomsNoImpervDissolve,clippingBoundary,"perviousGeomsNoImperv_clip")[0]
                areas["Total Pervious From Landcover Minus Building Footprints and Parking and Roads"]["value"]=self.calculateAcreAreaFromPolygons(perviousGeomsNoImperv_clip)
                print(areas)
        if perviousGeomsNoImperv_clip and treesLayerAll:
            arcpy.env.workspace = self.ProjectDatabase
            arcpy.env.overwriteOutput = True
            perviousTreesSoil = arcpy.Intersect_analysis([perviousLayer,treesLayerAll],"pervious_trees_intersect")[0]
            perviousTreesSoil = arcpy.Dissolve_management(perviousTreesSoil,"pervious_trees_soil")[0]
            areas["Pervious (Trees)"]["value"]=self.calculateAcreAreaFromPolygons(perviousTreesSoil)
            print(areas)
            treesGeometry = arcpy.CopyFeatures_management(treesFile,arcpy.Geometry())
            treesFileSR = arcpy.Describe(treesFile).spatialReference
            perviousGeoms = arcpy.CopyFeatures_management(perviousFile,arcpy.Geometry())
            perviousGeomsNoTrees = []
            for perviousGeom in perviousGeoms:
                for tree in treesGeometry:
                    tree_sr = tree.projectAs(perviousGeom.spatialReference)
                    perviousGeom = perviousGeom.difference(tree_sr)
                perviousGeomsNoTrees.append(perviousGeom)

            perviousNoTreesDissolve = arcpy.Dissolve_management(perviousGeomsNoTrees,"allPerviousNoTrees_dissolve")[0]
            perviousNoTrees_clip = arcpy.Clip_analysis(perviousNoTreesDissolve,clippingBoundary,"allPerviousNoTrees_clip")[0]
            areas["Pervious (No Trees)"]["value"]=self.calculateAcreAreaFromPolygons(perviousNoTrees_clip)

        self.writeValuesToAreasTable(areas)


    def exportNewToKML(self):

        currentClassFolder = os.path.dirname(os.path.realpath(__file__))
        mxd_report = arcpy.mapping.MapDocument(self.ProjectReportDocument)

        newName = arcpy.CreateUniqueName(self.ProjectName+"_kml.mxd", self.ProjectFolder)
        oldName = currentClassFolder + "\\for_kmz_export_legend.mxd"
        mxd_kml_pth = arcpy.Copy_management(oldName,newName,"MapDocument")[0]
        mxd_kml = arcpy.mapping.MapDocument(mxd_kml_pth)
        mxd_kml_df = arcpy.mapping.ListDataFrames(mxd_kml)[0]
        lyrs = arcpy.mapping.ListLayers(mxd_report)
        cats = ["Elevation","LandCover","Enviroatlas"]
        demFile = ""
        totalFound = 0
        resultsList = {}
        legendLayersToKeep = []
        for cat in cats:
            fileTypes = self.getFileTypesFromCat(cat)
            for ft,geo,ub in fileTypes:
                eflst = self.getFiles(ft)
                if len(eflst)==1:
                    if cat == "Elevation":
                        rastLyr = arcpy.MakeRasterLayer_management(eflst[0],"Digital Elevation Model")[0]
                        print rastLyr
                        arcpy.mapping.AddLayer(mxd_kml_df,rastLyr,"TOP")
                        legendLayersToKeep.append(rastLyr)
                        print legendLayersToKeep
                        #mxd_kml.save()
                    if cat == "LandCover" :
                        if geo == "Raster":
                            rastLyr = arcpy.MakeRasterLayer_management(eflst[0],"Landcover")[0]
                            print rastLyr
                            arcpy.mapping.AddLayer(mxd_kml_df,rastLyr,"TOP")
                            legendLayersToKeep.append(rastLyr)
                            print legendLayersToKeep
                            #mxd_kml.save()
                    if cat == "Enviroatlas":
                        if geo == "Raster":
                            rastLyr = arcpy.MakeRasterLayer_management(eflst[0],"Enviroatlas")[0]
                            print rastLyr
                            arcpy.mapping.AddLayer(mxd_kml_df,rastLyr,"TOP")
                            legendLayersToKeep.append(rastLyr)
                            print legendLayersToKeep
                            #mxd_kml.save()
        removeList = []
        for lyr in lyrs[::-1]:
            arcpy.mapping.AddLayer(mxd_kml_df,lyr,"TOP")
            ul= arcpy.mapping.ListLayers(mxd_kml,lyr.name)[0]
            if lyr.name == "FEMA Floodplain Layer (S_FLD_HAZ_AR)":
                legendLayersToKeep.append(ul)
            elif lyr.name == "Polygon Soils Layer with HYDGRP Field":
                legendLayersToKeep.append(ul)
            else:
                removeList.append(ul)


        legend = arcpy.mapping.ListLayoutElements(mxd_kml,"LEGEND_ELEMENT")[0]
        #print removeList

        for llyr in removeList:
            try:
                legend.removeItem(llyr)
            except:
                print llyr
        mxd_kml_df.name = self.ProjectName


        mxd_kml.save()
        del(mxd_kml)
        newNKMZ = arcpy.CreateUniqueName(self.ProjectName+"_kml.kmz", self.ProjectFolder)
        arcpy.MapToKML_conversion(mxd_kml_pth,self.ProjectName,newNKMZ)

